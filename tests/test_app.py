import sqlite3

import pytest
from openpyxl import load_workbook

import app_logic
import storage


def test_labor_calculation_uses_tier_rate():
    gross, full, partial = app_logic.calculate_labor_pay(2.5, "Skill")
    assert full == 1300.0
    assert partial == 325.0
    assert gross == 1625.0


def test_monthly_totals_separate_payroll():
    totals = app_logic.monthly_totals(
        [{"type": "material", "month": "2026-08", "amount": 1200}],
        [{"month": "2026-08", "net": 500}],
        [{"month": "2026-08", "price": 100}],
        "2026-08",
    )
    assert totals == {"materials": 1200.0, "construction": 0.0, "labor": 500.0, "payroll": 100.0, "total": 1800.0}


def test_excel_exports_have_expected_sheets():
    assert load_workbook("ailyn_project_ledger.xlsx", read_only=True).sheetnames == [
        "Transactions", "Payroll", "Monthly Summary", "Receipt Archive"
    ]
    assert load_workbook("materials_ledger.xlsx", read_only=True).sheetnames == ["Materials", "Monthly Summary"]
    assert load_workbook("labor_ledger.xlsx", read_only=True).sheetnames == ["Labor", "Monthly Summary"]


def test_sqlite_save_and_reset(monkeypatch, tmp_path):
    database = tmp_path / "test.db"
    monkeypatch.setattr(storage, "DB_FILE", str(database))
    monkeypatch.setattr(storage, "LEGACY_STATE_FILE", str(tmp_path / "missing.json"))
    monkeypatch.setattr(storage, "BACKUP_DIR", str(tmp_path / "backups"))
    state = storage.load_state()
    state["records"] = [{"id": "receipt-1", "amount": 99}]
    state["client_notes"] = [{"id": "note-1", "folder": "Client Approval", "text": "Approved", "created_at": "2026-08-27T10:00:00+08:00"}]
    storage.save_state(state)
    assert storage.load_state()["records"][0]["id"] == "receipt-1"
    assert storage.load_state()["client_notes"][0]["folder"] == "Client Approval"
    storage.save_state({**state, "records": []})
    assert storage.load_state()["records"] == []
    backup = storage.create_backup()
    assert sqlite3.connect(backup).execute("SELECT 1").fetchone() == (1,)


def test_scanner_photo_save_and_delete(monkeypatch, tmp_path):
    monkeypatch.setattr(storage, "APP_DIR", str(tmp_path))
    monkeypatch.setattr(storage, "SCANNER_PHOTO_DIR", str(tmp_path / "scanner"))

    relative_path = storage.save_scanner_photo(b"photo-bytes", "image/jpeg", "photo-1")
    assert relative_path == "scanner/photo-1.jpg"
    assert (tmp_path / "scanner" / "photo-1.jpg").read_bytes() == b"photo-bytes"
    assert storage.delete_scanner_photo(relative_path) is True
    assert not (tmp_path / "scanner" / "photo-1.jpg").exists()
