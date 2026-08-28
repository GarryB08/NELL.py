import os
import time
import base64
import ast
from io import BytesIO
import operator
import smtplib
import shutil
import tempfile
import hashlib
import hmac
import uuid
import re
from datetime import datetime
from email.message import EmailMessage
from zoneinfo import ZoneInfo
import streamlit as st
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageOps
try:
    import pytesseract
except ModuleNotFoundError:
    pytesseract = None
TESSERACT_NOT_FOUND_ERROR = getattr(pytesseract, "TesseractNotFoundError", RuntimeError)
from storage import BACKUP_DIR, DB_FILE, create_backup, delete_scanner_photo, history_count, load_state as load_sqlite_state
from storage import restore_backup, save_scanner_photo, save_state as save_sqlite_state
from app_logic import FULL_DAY_RATES, TIER_TABLE, calculate_labor_pay, get_partial_rate

APP_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_FILE = os.path.join(APP_DIR, "app_state.json")
EXCEL_FILE = os.path.join(APP_DIR, "ailyn_project_ledger.xlsx")
MATERIALS_EXCEL_FILE = os.path.join(APP_DIR, "materials_ledger.xlsx")
LABOR_EXCEL_FILE = os.path.join(APP_DIR, "labor_ledger.xlsx")
PHILIPPINES_TZ = ZoneInfo("Asia/Manila")


def manila_now():
    return datetime.now(PHILIPPINES_TZ)


def scan_photo_text(photo):
    """Read text from a camera photo using the local OCR engine."""
    if pytesseract is None:
        raise RuntimeError("OCR is unavailable. Install the pytesseract package to scan text.")
    image = Image.open(BytesIO(photo.getvalue()))
    return pytesseract.image_to_string(image).strip()


def parse_scanned_receipt(text):
    """Extract conservative material-entry suggestions from OCR text."""
    amount_pattern = r"(?:PHP|₱)\s*([0-9][0-9,]*(?:\.\d{1,2})?)"
    currency_amounts = [float(value.replace(",", "")) for value in re.findall(amount_pattern, text, re.IGNORECASE)]
    all_amounts = [float(value.replace(",", "")) for value in re.findall(r"\b[0-9][0-9,]*\.\d{2}\b", text)]
    amounts = currency_amounts or all_amounts
    quantity_match = re.search(r"\b(?:qty|quantity)\s*[:x-]?\s*(\d+)\b|\b(\d+)\s*(?:pcs?|pieces?|units?|x)\b", text, re.IGNORECASE)
    quantity = int(next(value for value in quantity_match.groups() if value)) if quantity_match else 1
    delivery_match = re.search(r"delivery\D+([0-9][0-9,]*(?:\.\d{1,2})?)", text, re.IGNORECASE)
    delivery = float(delivery_match.group(1).replace(",", "")) if delivery_match else 0.0
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    name = next((line for line in lines if not re.search(r"(?:php|₱|total|qty|quantity|delivery|receipt|invoice|date)", line, re.IGNORECASE)), "")
    total = amounts[-1] if amounts else 0.0
    price = total / quantity if quantity > 0 else total
    return {"name": name[:120], "price": price, "qty": quantity, "delivery": delivery}


def normalize_photo_bytes(photo_bytes, mime_type="image/jpeg"):
    """Apply camera EXIF orientation and return display-ready image bytes."""
    image = ImageOps.exif_transpose(Image.open(BytesIO(photo_bytes))).convert("RGB")
    output = BytesIO()
    image.save(output, format="JPEG", quality=90, optimize=True)
    return output.getvalue(), "image/jpeg"


def searchable_records(state):
    records = []
    for category, key in (("Financial", "records"), ("Labor", "labor_records"), ("Payroll", "payroll_expenses")):
        for record in state.get(key, []):
            row_category = "Expense" if category == "Financial" and record.get("type") == "expense" else "Material" if category == "Financial" else category
            records.append({"category": row_category, **record})
    return records


def find_duplicate_records(records):
    groups = {}
    for record in records:
        signature = (
            record.get("type", record.get("category", "")),
            str(record.get("name", record.get("item", record.get("description", "")))).strip().lower(),
            round(float(record.get("amount", record.get("price", record.get("net", 0))) or 0), 2),
            record.get("date", record.get("month", "")),
        )
        groups.setdefault(signature, []).append(record)
    return [group for group in groups.values() if len(group) > 1]


# --- PERSISTENCE HELPERS
PERSISTENT_KEYS = [
    "records",
    "labor_records",
    "payroll_expenses",
    "planner_tasks",
    "budget",
    "budget_history",
    "remaining_money",
    "view",
    "receipt_archive",
    "project",
    "scanner_photos",
    "dark_mode",
    "client_notes",
    "app_settings",
]

def load_state():
    return load_sqlite_state()


def month_key(record):
    recorded_at = record.get("recorded_at")
    if recorded_at:
        try:
            return datetime.fromisoformat(recorded_at).strftime("%Y-%m")
        except ValueError:
            pass
    try:
        return datetime.strptime(record.get("date", ""), "%b %d, %Y").strftime("%Y-%m")
    except ValueError:
        return manila_now().strftime("%Y-%m")


def write_excel(state, archive_entry=None):
    """Write the current app data to a durable workbook with monthly totals."""
    workbook = Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "Transactions"
    transaction_headers = ["Date", "Month", "Category", "Description", "Supplier", "Invoice", "Payment Method", "Sender", "Qty", "Unit Price", "Delivery", "Amount"]
    transactions_sheet.append(transaction_headers)
    for record in state.get("records", []):
        transactions_sheet.append([
            record.get("date", ""), month_key(record), record.get("type", "").title(),
            record.get("name", ""), record.get("supplier", ""), record.get("invoice_number", ""),
            record.get("payment_method", ""), record.get("sender", ""), record.get("qty", 1),
            float(record.get("price", 0)), float(record.get("delivery", 0)), float(record.get("amount", 0)),
        ])

    payroll_sheet = workbook.create_sheet("Payroll")
    payroll_sheet.append(["Date", "Month", "Type", "Worker ID", "Worker / Description", "Role", "Pay Period", "Days", "Gross Pay", "Cash Advance", "Net / Amount"])
    for record in state.get("labor_records", []):
        payroll_sheet.append([
            record.get("date", ""), month_key(record), "Labor", record.get("worker_id", ""), record.get("name", ""),
            record.get("role", ""), record.get("pay_period", ""), float(record.get("days", 0)), float(record.get("gross_pay", 0)),
            float(record.get("ca", 0)), float(record.get("net", 0)),
        ])
    for record in state.get("payroll_expenses", []):
        payroll_sheet.append([
            record.get("date", ""), month_key(record), "Payroll Expense", record.get("item", ""),
            "", "", "", "", float(record.get("price", 0)),
        ])

    summary = workbook.create_sheet("Monthly Summary")
    summary.append(["Month", "Materials", "Construction Expenses", "Excess Money", "Labor", "Payroll Expenses", "Total Spent"])
    months = {month_key(record) for record in state.get("records", [])}
    months.update(month_key(record) for record in state.get("labor_records", []))
    months.update(month_key(record) for record in state.get("payroll_expenses", []))
    for month in sorted(months, reverse=True):
        materials = sum(float(r.get("amount", 0)) for r in state.get("records", []) if month_key(r) == month and r.get("type") == "material")
        construction = sum(float(r.get("amount", 0)) for r in state.get("records", []) if month_key(r) == month and r.get("type") == "expense")
        excess = sum(float(r.get("amount", 0)) for r in state.get("records", []) if month_key(r) == month and r.get("type") == "excess")
        labor = sum(float(r.get("net", 0)) for r in state.get("labor_records", []) if month_key(r) == month)
        payroll = sum(float(r.get("price", 0)) for r in state.get("payroll_expenses", []) if month_key(r) == month)
        summary.append([month, materials, construction, excess, labor, payroll, materials + construction + labor + payroll])

    archive = workbook.create_sheet("Receipt Archive")
    archive.append(["Receipt ID", "Saved At", "Report Type", "Title", "HTML File"])
    for entry in state.get("receipt_archive", []):
        archive.append([entry.get("id", ""), entry.get("saved_at", ""), entry.get("report_type", ""), entry.get("title", ""), entry.get("file", "")])
    if archive_entry:
            archive.append([archive_entry["id"], archive_entry["saved_at"], archive_entry["report_type"], archive_entry["title"], archive_entry["file"]])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1B5E20")
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
            sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(EXCEL_FILE)


def write_separate_excel_files(state):
    """Write focused workbooks so materials and labor can be shared independently."""
    materials_book = Workbook()
    materials_sheet = materials_book.active
    materials_sheet.title = "Materials"
    materials_sheet.append(["Date", "Month", "Description", "Supplier", "Invoice", "Payment Method", "Sender", "Qty", "Unit Price", "Delivery", "Amount"])
    for record in state.get("records", []):
        if record.get("type") == "material":
            materials_sheet.append([
                record.get("date", ""), month_key(record), record.get("name", ""), record.get("supplier", ""),
                record.get("invoice_number", ""), record.get("payment_method", ""), record.get("sender", ""),
                record.get("qty", 1), float(record.get("price", 0)), float(record.get("delivery", 0)),
                float(record.get("amount", 0)),
            ])

    labor_book = Workbook()
    labor_sheet = labor_book.active
    labor_sheet.title = "Labor"
    labor_sheet.append(["Date", "Month", "Worker ID", "Worker", "Role", "Pay Period", "Days", "Gross Pay", "Cash Advance", "Net Pay"])
    for record in state.get("labor_records", []):
        labor_sheet.append([
            record.get("date", ""), month_key(record), record.get("worker_id", ""), record.get("name", ""), record.get("role", ""), record.get("pay_period", ""),
            float(record.get("days", 0)), float(record.get("gross_pay", 0)), float(record.get("ca", 0)),
            float(record.get("net", 0)),
        ])

    for workbook, path in ((materials_book, MATERIALS_EXCEL_FILE), (labor_book, LABOR_EXCEL_FILE)):
        summary = workbook.create_sheet("Monthly Summary")
        summary.append(["Month", "Total"])
        source_records = state.get("records", []) if path == MATERIALS_EXCEL_FILE else state.get("labor_records", [])
        months = sorted({month_key(record) for record in source_records}, reverse=True)
        for month in months:
            total = sum(
                float(record.get("amount", 0) if path == MATERIALS_EXCEL_FILE else record.get("net", 0))
                for record in source_records if month_key(record) == month
            )
            summary.append([month, total])
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1B5E20")
            for column in sheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 32)
                sheet.column_dimensions[column[0].column_letter].width = width
        workbook.save(path)


def save_state(state):
    data = save_sqlite_state(state)
    write_excel(state)
    write_separate_excel_files(state)

APP_VERSION = "AILYN HOUSE"
APP_NAME = "AILYN HOUSE | Ailyn House Project"

# ================================================================
# Construction/Materials, Payroll, and Schedule are intentionally unified.
# ================================================================

RECEIVER_EMAIL = "garryboypepito2004@gmail.com"
RECEIVER_AILYN = "ailyn_peps0678@yahoo.com"
AILYN_LOGO_DATA = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAJsAAAC5CAYAAAA/BU2xAAANMklEQVR4nO3de1gU9RrA8XcEr5kmnvB21DI1zSOleclQIw0yj2bFelBKCCQJD3hBQVGCRRINEJRFyAviXbSFQBHjYooJKpgim2IqykXIW2haltbj7/zRc56HR0V22Zn3NzP7fv51mHmd/Trs7M6MAmMMCMHQjPcAxHJQbAQNxUbQUGwEDcVG0FBsBA3FRtBQbAQNxUbQUGwEDcVG0FBsBA3FRtBQbAQNxUbQUGwEDcVG0FBsnHxz6juLu0SaYuNgy3cZLCYrmfcY6Cg2ZDFZyeyTdcG8x+DCmvcAliQoZQWLzdrIewxuKDYkXmuD2dbDGbzH4IpiQzAp2odllx7mPQZ3FJvE7EOnsO8vneY9hizQCYJEqm7Usj5znSi0eujIJoETl86wsUvd4Pf793iPIit0ZBNZnqGQvR7qQqE9BsUmop1HstiEKG/eY8gWxSaSxNwdzD1xAe8xZI3es4kgPC2BLU1P5D2G7FFsZpq7JYIl5u7gPYYiUGxmcE9cwHYeyeI9hmJQbE1w/XYd81gTBHmGQt6jKArFZqKztRfZx4kLoaSyjPcoikOxmaDw3Ak2LSEQauqu8h5FkSg2I+05cYB9GD8P7v/1J+9RFIs+ZzPCxvw0NnnlLArNTHRka8SKvRvY4p2xvMdQBTqyNYJCEw/FRtBQbAQNxUbQUGwEDcVG0FBsBA3FRtBQbAQNxUbQUGwEDcVG0FBsBA3FRtBQbASNImMLT0uwuOfRqoHiYgvYFkmhKZSiYpux/jOmy97CewzSRIq4LPynW9fZ3M0RkH48j/coxAyyj62sppz5b10GB04f4z0KMZOsYzt6voTN27oc6OmN6iDb2LJLD7N5W5fDhSuVvEchIpFlbLuO7mPzti6H67freI9CRCS72NYf+IrN27oc7v15n/coRGSyio1uCFY32cQW8tUqFrlnPe8xiIRkEdvsTUvZmv0pvMcgEuMe28dfLmQphXt5j0EQcI3tgxhfllWSz3MEgojLd6PlV6vY2KXuFJqFQT+ynag4w7zXfQaG6nPYmyacocZ24Mwx5r0+BKpu1GJulsgEWmxfF+cyn6RQuHX3DtYmicygxLYxP435bNACY3TdoyWT/ARh5b5N7NOkUAoNiaH6HDtUVizLnS1pbGGp8WzhjmgpN0HqST+ex4YuduY9RoMk+zU6d3MES8xT9v/pNDHS+9EjBAMAAX+Wxijhe2VJYvNcs4htL9gjxarRjAhxYScrzjz6BzIMzSdJy5LzU3mP0SjRY3OO9WN7Tx4Ue7VoLtddYW8t9YCK65cbX1gGRzmnZZ7sUFkx3yGMJFpsl+uuMPfEBVDw4wmxVomuqNzAPoj5L9y4c5P3KI0yVJ9jk6J9oPbmNd6jGE202N5bMRN+qD4v1urQbTz0NfPfEgF37/1h/A9xOqqlH89jU+Lm8tm4GUSLTcmhLd4Zw2L3bYIHDx40fSWMAQjS16eEE4GGcL/EiLcpurksvViE+1ERQlPKiUBDLDa2ogulzG9zOJyqOCv+yiU4cVDSiUBDLDK2zYfSWeD2KLh197Y0GxAxNCWeCDTE4mJblBLDYrKSeY9hFKWeCDTEYmI7WVHGwtNWg1Iu2FTyiUBDLCK2HYV7Wag+TjHX0Sn9RKAhqo9Nq9ex5bvX8h7DaGo4EWiIamM7ffkCC09brZjHbKnpRKAhqowttSiHfbZrJVy8Vs17lAaVXb4A5Ver2AudeghqOxFoiOpiW5r+JQtPW817jEZdv3MTNCtnwcs9+7OUwkze46BQTWznr1Sw8LQE2HV0H+9RjFZWUw5lNeW8x0CjitgyTx5ki1Ni4MefLkm2jT6de8J5elacWRT1AOfHid67gWli/SQLrW+X52Cb7wowRGYKgRO9JNmGpVBsbFU3atn0tYtZsIQffC6a5A2lX+wRnIc5CQAASybPFoo+18P4V96QbJtqpshfo+VXq+CDWF/JLmt699UxoNX4wUvdej/yLaddjxeFNP942F6wh4XqdVD980+SzKBGgli32LVyG4h0+5gAf19WIb5ett0h1NkXXEaMN/qrdDl+15oTtAFG9x8qu7slFPhrVJrQAid6wZnoLMGU0AAAIqb4CwVhKeBkN1KSudREgbGJa/wrb0DR53pYMnl2k48Erz4/QNg9P1FY6xUOXTrYijmeqlhsbN07doEN3hGQ5h8v2PV4UZRfOW6j3xMurdovzBrnJsbqVMciY/Mf7wHnY3MEV/uJkryviXQNEA6FbIMxA16TYvWKZVGxOdmNhIKwFIiY4i/5m+dhve2ErAXrhATPUHi2nY00G5HdKcCTWURsXTrYwlqvcNg9P1F49fkBqC+Rp4NGqI7PF2Y6uoq/clk+PqZhqo9t1jg3uLRqv+A2+j2ux4GYaUHCt8GbYFS/IRKsXRmHONXGNmbAa3AoZBtEugbI5pV4ve9gIXdRshDnHgw2T7UXcc3KOMSpLrZn29lAgmcoZC1YJwzrbSeb0OqbMdZFqE08LMwY68J7FFT8YpPgH+NMR1eojs8XPB00sozsYXHuwUJO0AYY0WeQqOvNOL5f1PWJRYFfVz1qVL8hEKbxg9f7DlZEZI+TkLudafU6uP37r6Ksz330+7DGa4ms9oeiY7N5qj1oNX4wY6yLrHZqU1375WcWmqqD5IPi3Fk1tNdAiHELgqG9Bspi/yg2thljXSDOPVgWO1Fs354+ykL1OiguLzV7XW1btYHlU+eD15uTue8rxcU2os8gCNP4yfKqBrHFfbOZafU6uHvfhMd4NcDrzckQ7xHCdZ8hxCYAA2b2J0HtWrcFrcYPZjq6qj6y+mpvXmOh+jjY8l2G2esa1tsO4tyD4ZWe/bnsQ0Uc2TwcnCHM2Q9s23e0qNDqyyk9zLR6HZx43HN+TdCudVuI+jAQ3Ee/j74vZR3b0BfsIEzjB2MGvGaxkT0sJiuZafU6uP/Xn2atx+etqRDrtgh1v8oytjYtWoFW4wezxrlRZI9RdaOWhep1sMPM+01H9BkE8R4hMOCfj17+LgXZxTZt1CQI08yCrh1sKbRGZJXkM61eB6VVPzZ5Hc+0eRpipgWBVJdb1SdKbJfrrrDecxzNWsfg514CrcYPnOxGUmQm+mL3OhaWGg8PWNOfCTxr3DSIdA2UdN+bHdu+kkPMd+MSqKm72uR1REzxB//xHpL8RU9VnpXdt9Qv9+wn+t/14rVqpk3Vwa4jTX8iwOh+QyHeIwT6dnlOktfCrNiWZaxhYanxZg3Qqf0/oFJ3QLJ/UW4JgcyURzIETJgOLZu3NHr5ovJSyCk9bPTyu+cnSnb0zik9zN6N9gHHgfaQayho0jps2raHVe7BMHn4ONFnbNJ9o9d++Zn5bQyHjO/l+YWvOcL/M8eknRyfvZWZEhuGgAnTYd0n4RCVmQSrc7aZ9LN1v/4C01YHwKnKs+xzE/dFY0y+6iPXUMBGhbmqMjQ16fzMs8KKjxYKf2w2CIsmeUNzK9OOK9GZSTAhyptVXK8R7W2ISbFFZSaxiVGfQqXJjwul9/wYnOxGCn9sNggPf5UX4uwr3Ek+KSyfOh86Pt3B6PXlGQphpHYqpB/PEyU4o3P/aPV8pj+WbfSKh/T6FzjZ2UO3Dp3B2soKrIRmYGVlDdZWVmDdzAqsrazBupkVtGnZukmDE9PNecddmPOOO2w4qGdRe5LgkhH/GdyNOzdhStxcCJrkzUKdfc06ajQa28Ezx5jvxnC4YOTjomY6ukLAhOnQhT4nky1PB43g6aAB/bFsFp2ZBCWVZY3+zLKMNVBSWcbiPw6BbjadmvTaPjG2lfs2saySfOjawRa6NnKnt33fwWBu+QSXZvjbgmb425BrKGBRmUmNLv/bvd/BY00QzB7nxv49yMHk1/qJsf3/sEvUzXGgveA40F7y7ajuhhciX6J9NypXaUU5zJQ70kf1G2LSr4fam9dY+dUqo5dvYd0chvd+2SLfbijyYYCm6GbTSdIXt2sHW6Gx97P1nbh0Wt3/up+Afo0SNBQbQUOxETQUG0Gj+hOEovJSyDUUGv2mPPh9H5NOJorLS1l2qfGX8zj0H2rK6lUFNbbkg6noZ2K7jmRB8cUfjF7e5qn2rHUL469n+/b0Ufjq2DdGL3/rt4/g/JVK9P3g4eDM/eMW1M/ZArZFMl32FrTtkb+1at4SbiUd5x4bvWcjaCg2ggb1PVvbVm0gaJI35iYh4/h+OFNzwejl/cd7QMvmLYxevrjcAHk/FBq9vMuI8dDLtrvRy6sJamw8LkEqv1rFenfuYfTymuFvQysTThC6d+wMbVq2Mnr5D+0nWuztiqr/Ip7IB71nI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoKGYiNoKDaChmIjaCg2goZiI2goNoLmfyASapdvGeRmAAAAAElFTkSuQmCC"
SENDER_EMAIL = os.getenv("AILYN_SENDER_EMAIL", "")
SENDER_PASSWORD = os.getenv("AILYN_SENDER_PASSWORD", "")
ADMIN_PASSWORD = os.getenv("AILYN_ADMIN_PASSWORD", "")
UPDATE_SIGNING_KEY = os.getenv("AILYN_UPDATE_SIGNING_KEY", "")
LOGIN_PASSWORD = os.getenv("AILYN_LOGIN_PASSWORD", "")

st.set_page_config(
    page_title=APP_NAME,
    page_icon="🅰️",
    layout="wide",
)

if LOGIN_PASSWORD and not st.session_state.get("authenticated"):
    st.title("Ailyn House Project")
    st.caption("Sign in to access this project workspace.")
    login_password = st.text_input("Workspace password", type="password")
    if st.button("SIGN IN", use_container_width=True):
        if hmac.compare_digest(login_password, LOGIN_PASSWORD):
            st.session_state.authenticated = True
            st.rerun()
        st.error("Invalid workspace password.")
    st.stop()

# Load persisted state safely
state_data = load_state()
for key in PERSISTENT_KEYS:
    if key in state_data and key not in st.session_state:
        st.session_state[key] = state_data[key]

if "records" not in st.session_state:
    st.session_state.records = []
if "labor_records" not in st.session_state:
    st.session_state.labor_records = []
if "payroll_expenses" not in st.session_state:
    st.session_state.payroll_expenses = []
if "receipt_archive" not in st.session_state:
    st.session_state.receipt_archive = []
if "scanner_photos" not in st.session_state:
    st.session_state.scanner_photos = []
if "project" not in st.session_state:
    st.session_state.project = {
        "name": "Ailyn House Project", "client": "", "address": "", "manager": "",
        "status": "Active", "target_date": "",
    }
if "planner_tasks" not in st.session_state:
    st.session_state.planner_tasks = []
if "budget" not in st.session_state:
    st.session_state.budget = 0.0
if "budget_history" not in st.session_state:
    st.session_state.budget_history = []
if "remaining_money" not in st.session_state:
    st.session_state.remaining_money = 0.0
if "view" not in st.session_state:
    st.session_state.view = "home"
if st.session_state.view not in {
    "home", "payroll_dashboard", "planner_input", "planner_output", "material",
    "expense", "excess", "ledger", "add_labor", "add_payroll_expense",
    "payroll_remaining", "payroll_ledger", "export", "payroll_export",
    "receipt_archive", "photo_scanner", "project_tools", "settings", "update",
}:
    st.session_state.view = "home"
if "selected_role" not in st.session_state:
    st.session_state.selected_role = "Labor"
if "editing_record_id" not in st.session_state:
    st.session_state.editing_record_id = None
if "editing_labor_index" not in st.session_state:
    st.session_state.editing_labor_index = None
if "editing_payroll_expense_index" not in st.session_state:
    st.session_state.editing_payroll_expense_index = None
if "scanner_input_version" not in st.session_state:
    st.session_state.scanner_input_version = 0
if "scanner_actions_open" not in st.session_state:
    st.session_state.scanner_actions_open = False
if "scanner_open" not in st.session_state:
    st.session_state.scanner_open = False
if "scanner_flash_mode" not in st.session_state:
    st.session_state.scanner_flash_mode = "Auto"
if "scanner_camera_mode" not in st.session_state:
    st.session_state.scanner_camera_mode = "Back camera"
if "client_notes" not in st.session_state:
    st.session_state.client_notes = []
if "app_settings" not in st.session_state:
    st.session_state.app_settings = {
        "display_name": "",
        "email": "",
        "client_mode": False,
        "email_notifications": True,
        "budget_alerts": True,
        "date_format": "%Y-%m-%d",
    }
else:
    st.session_state.app_settings = {
        "display_name": "",
        "email": "",
        "client_mode": False,
        "email_notifications": True,
        "budget_alerts": True,
        "date_format": "%Y-%m-%d",
        **st.session_state.app_settings,
    }
if "dark_mode" not in st.session_state:
    st.session_state.dark_mode = False
if not os.path.exists(EXCEL_FILE):
    write_excel(st.session_state)
if not os.path.exists(MATERIALS_EXCEL_FILE) or not os.path.exists(LABOR_EXCEL_FILE):
    write_separate_excel_files(st.session_state)


def set_view(v):
    st.session_state.view = v
    persist_state()
    st.rerun()


def project_settings_dialog():
    project = st.session_state.project
    with st.form("project_settings_form"):
        name = st.text_input("Project name", value=project.get("name", "Ailyn House Project"))
        client = st.text_input("Client name", value=project.get("client", ""))
        address = st.text_input("Site address", value=project.get("address", ""))
        manager = st.text_input("Project manager", value=project.get("manager", ""))
        status = st.selectbox("Project status", ["Planning", "Active", "On Hold", "Completed"], index=["Planning", "Active", "On Hold", "Completed"].index(project.get("status", "Active")))
        target_date = st.date_input("Target completion", value=datetime.fromisoformat(project["target_date"]).date() if project.get("target_date") else manila_now().date())
        if st.form_submit_button("SAVE PROJECT DETAILS", use_container_width=True):
            if not name.strip():
                st.error("Project name is required.")
            else:
                st.session_state.project = {"name": name.strip(), "client": client.strip(), "address": address.strip(), "manager": manager.strip(), "status": status, "target_date": target_date.isoformat()}
                persist_state()
                st.success("Project details saved.")


@st.dialog("Notes")
def notes_dialog():
    st.caption("Project notes, approvals, and photo comments")
    folders = ["Site Notes", "Client Approval", "Photo Comments", "Change Requests"]
    folders += [folder for folder in sorted({note.get("folder", "Site Notes") for note in st.session_state.client_notes}) if folder not in folders]
    folder = st.selectbox("Folder", folders, key="popup_note_folder")
    new_folder = st.text_input("New folder", key="popup_new_note_folder", placeholder="Optional folder name")
    text = st.text_area("Note", key="popup_note_text", height=80, placeholder="Write a note for the project...")
    if st.button("ADD NOTE", use_container_width=True, key="popup_add_note"):
        if text.strip():
            st.session_state.client_notes.insert(0, {"id": str(uuid.uuid4()), "folder": new_folder.strip() or folder, "text": text.strip(), "created_at": manila_now().isoformat()})
            persist_state()
            st.rerun()
        st.warning("Write a note before saving.")
    if st.session_state.client_notes:
        st.markdown("#### RECENT NOTES")
        st.dataframe([
            {"Folder": note.get("folder", "Site Notes"), "Date": note.get("created_at", "").replace("T", " ")[:16], "Note": note.get("text", "")}
            for note in st.session_state.client_notes[:8]
        ], use_container_width=True, hide_index=True)
        note_to_delete = st.selectbox("Delete note", ["Choose a note"] + [f"{note.get('created_at', '')[:16]} | {note.get('text', '')[:45]}" for note in st.session_state.client_notes[:8]], key="popup_delete_note_select")
        if note_to_delete != "Choose a note" and st.button("DELETE SELECTED NOTE", use_container_width=True, key="popup_delete_note"):
            selected_index = [f"{note.get('created_at', '')[:16]} | {note.get('text', '')[:45]}" for note in st.session_state.client_notes[:8]].index(note_to_delete)
            st.session_state.client_notes.pop(selected_index)
            persist_state()
            st.rerun()
    else:
        st.info("No notes yet.")


@st.dialog("Settings")
def settings_dialog():
    settings = st.session_state.app_settings
    st.caption("Quick account and project preferences")
    with st.form("quick_settings_form"):
        display_name = st.text_input("Display name", value=settings.get("display_name", ""))
        email = st.text_input("Email", value=settings.get("email", ""))
        dark_mode = st.checkbox("Dark mode", value=bool(st.session_state.dark_mode))
        budget_alerts = st.checkbox("Budget alerts", value=bool(settings.get("budget_alerts", True)))
        email_notifications = st.checkbox("Email notifications", value=bool(settings.get("email_notifications", True)))
        if st.form_submit_button("SAVE SETTINGS", use_container_width=True):
            if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
                st.error("Enter a valid email address.")
            else:
                settings.update({"display_name": display_name.strip(), "email": email.strip(), "budget_alerts": budget_alerts, "email_notifications": email_notifications})
                st.session_state.dark_mode = dark_mode
                persist_state()
                st.success("Settings saved.")
    st.markdown("#### ACCOUNT ACCESS")
    st.success("Workspace password enabled.") if LOGIN_PASSWORD else st.info("Workspace password is disabled.")
    if LOGIN_PASSWORD and st.button("SIGN OUT", use_container_width=True, key="popup_sign_out"):
        st.session_state.authenticated = False
        st.rerun()


def total_materials():
    return sum(r["amount"] for r in st.session_state.records if r["type"] == "material")


def total_expenses():
    return sum(r["amount"] for r in st.session_state.records if r["type"] == "expense")


def total_excess():
    return sum(r["amount"] for r in st.session_state.records if r["type"] == "excess")


def get_total():
    return total_materials() + total_expenses()


def get_balance():
    return float(st.session_state.budget) - total_excess() - get_total()


def monthly_spend(month=None):
    month = month or manila_now().strftime("%Y-%m")
    construction = sum(
        float(record.get("amount", 0)) for record in st.session_state.records
        if record.get("type") in {"material", "expense"} and month_key(record) == month
    )
    labor = sum(float(record.get("net", 0)) for record in st.session_state.labor_records if month_key(record) == month)
    payroll_expenses = sum(float(record.get("price", 0)) for record in st.session_state.payroll_expenses if month_key(record) == month)
    return construction + labor + payroll_expenses


def monthly_construction_spend(month=None):
    month = month or manila_now().strftime("%Y-%m")
    return sum(
        float(record.get("amount", 0)) for record in st.session_state.records
        if record.get("type") in {"material", "expense"} and month_key(record) == month
    )


# ================================================================
# COMBINED MAIN RECEIPTS — FINANCIAL REPORT + PAYROLL REPORT
# Copied from the MAIN receipt implementation.
# ================================================================
def save_report_html(report_type, html_content, title="Receipt"):
    folder = os.path.join(APP_DIR, "archive", report_type)
    os.makedirs(folder, exist_ok=True)
    safe_title = re.sub(r"[^A-Za-z0-9._-]+", "_", title).strip("._") or "receipt"
    filename = os.path.join(folder, f"{safe_title}_{int(time.time())}.html")
    with open(filename, "w", encoding="utf-8") as f:
        f.write(html_content)
    archive_entry = {
        "id": str(uuid.uuid4()),
        "saved_at": manila_now().isoformat(),
        "report_type": report_type,
        "title": title,
        "file": os.path.relpath(filename, APP_DIR),
    }
    st.session_state.receipt_archive.append(archive_entry)
    persist_state()
    return filename


def list_saved_reports(report_type):
    folder = os.path.join(APP_DIR, "archive", report_type)
    if not os.path.exists(folder):
        return []
    from pathlib import Path
    return list(Path(folder).glob("*.html"))


def delete_report_file(path):
    if os.path.exists(path):
        os.remove(path)
    relative_path = os.path.relpath(path, APP_DIR)
    if "receipt_archive" in st.session_state:
        st.session_state.receipt_archive = [
            entry for entry in st.session_state.receipt_archive
            if entry.get("file") != relative_path
        ]
        persist_state()


def clear_saved_reports():
    for report_type in ("construction", "payroll"):
        for report_path in list_saved_reports(report_type):
            delete_report_file(report_path)
    st.session_state.receipt_archive = []
    persist_state()


def receipt_preview_height(item_count, row_height=58, base_height=560):
    """Reserve enough preview space for every report row and its totals."""
    return max(1, max(base_height, base_height + item_count * row_height) - 4)


def build_html_report(records, budget, custom_title="INVENTORY RECEIPT"):
    material_and_expense_records = [r for r in records if r["type"] in ["material", "expense"]]
    excess_records = [r for r in records if r["type"] == "excess"]
    material_total = sum(r["amount"] for r in material_and_expense_records)
    excess_total = sum(r["amount"] for r in excess_records)
    remaining_balance = get_balance()
    date_now = manila_now().strftime("%B %d, %Y")
    sobra_amount = 0.0
    kulang_amount = 0.0
    if remaining_balance > 0:
        sobra_amount = remaining_balance
    elif remaining_balance < 0:
        kulang_amount = abs(remaining_balance)
    balance_color = "#ffffff" if budget <= 0 else ("#e57373" if remaining_balance < 0 else "#a5d6a7")

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css?family=Inter:wght@400;600;700&display=swap');
body {{ font-family: 'Inter', sans-serif; background-color: #f0f4f0; margin: 0; padding: 20px; color: #333; }}
.receipt-container {{ max-width: 1000px; margin: auto; background: #fff; padding: 30px; border-radius: 4px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); border-top: 10px solid #1b5e20; }}
.header {{ display: flex; flex-wrap: wrap; justify-content: space-between; align-items: flex-start; margin-bottom: 30px; border-bottom: 2px solid #f0f0f0; padding-bottom: 15px; }}
.company-info h1 {{ color: #1b5e20; margin: 0; font-size: 24px; letter-spacing: -1px; }}
.company-info p {{ margin: 4px 0; font-size: 12px; color: #666; }}
.receipt-meta {{ text-align: left; margin-top: 10px; }}
@media (min-width: 768px) {{ .receipt-meta {{ text-align: right; margin-top: 0; }} }}
.receipt-meta h2 {{ margin: 0; font-size: 16px; text-transform: uppercase; color: #1b5e20; }}
.receipt-meta p {{ margin: 4px 0; font-size: 12px; font-weight: bold; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 12px; }}
th {{ background-color: #1b5e20; color: #ffffff; text-align: left; padding: 10px; text-transform: uppercase; letter-spacing: 1px; }}
td {{ padding: 10px 8px; border-bottom: 1px solid #f0f0f0; }}
.qty-col, .desccol, .pricecol, .deliverycol, .totalcol {{ text-align: left; }}
td.desccol {{ font-weight: 700; color: #333333; }}
th.desccol {{ color: #ffffff; }}
.summary-container {{ display: flex; justify-content: flex-end; }}
.summary-table {{ width: 100%; }}
@media (min-width: 768px) {{ .summary-table {{ width: 420px; }} }}
.grand-total {{ background: #013220; color: white; padding: 20px; border-radius: 4px; margin-top: 15px; }}
.balance-info {{ font-size: 13px; line-height: 1.8; }}
.balance-row {{ display: flex; justify-content: space-between; }}
.material-row {{ font-size: 18px; font-weight: bold; }}
.final-balance-row {{ display: flex; justify-content: space-between; border-top: 1px dashed rgba(255,255,255,0.4); margin-top: 8px; padding-top: 8px; font-size: 18px; font-weight: bold; }}
.footer {{ margin-top: 30px; text-align: center; font-size: 9px; color: #aaa; text-transform: uppercase; letter-spacing: 1px; }}
.save-btn-container {{ text-align: center; margin-bottom: 25px; }}
.save-img-btn {{ background-color: #1b5e20; color: white; border: none; padding: 12px 24px; font-size: 14px; font-weight: bold; border-radius: 6px; cursor: pointer; box-shadow: 0 4px 10px rgba(0,0,0,0.15); }}
.save-img-btn:hover {{ background-color:#013220; }}
@media print {{ .save-btn-container {{ display: none; }} }}
</style>
</head>
<body>
<div class="save-btn-container">
<button class="save-img-btn" onclick="saveAsImage()">SEE PHOTO & DOWNLOAD IMAGE (Phone & Laptop)</button>
</div>
<div class="receipt-container" id="receiptContent">
<div class="header">
<div class="company-info">
<h1>AILYN HOUSE PROJECT</h1>
<p>Official Material & Expense Inventory</p>
<p>Management System {APP_VERSION}</p>
<p>Backup Receiver: <i>{RECEIVER_AILYN}</i></p>
</div>
<div class="receipt-meta">
<h2>{custom_title}</h2>
<p>Date: {date_now}</p>
</div>
</div>
<table>
<thead>
<tr>
<th>Date</th>
<th class="qty-col">Qty</th>
<th class="desccol">Description</th>
<th class="pricecol">Unit Price</th>
<th class="deliverycol">Delivery</th>
<th class="totalcol">Total</th>
</tr>
</thead>
<tbody>"""

    for r in material_and_expense_records:
        html += f"""
<tr>
<td>{r['date']}</td>
<td class="qty-col">{r['qty']}</td>
<td class="desccol">{r['name']}</td>
<td class="pricecol">{float(r.get('price', r['amount'])):,.2f}</td>
<td class="deliverycol">{float(r['delivery']):,.2f}</td>
<td class="totalcol">PHP {float(r['amount']):,.2f}</td>
</tr>"""

    html += f"""
</tbody>
</table>
<div class="summary-container">
<div class="summary-table">
<div class="grand-total">
<div class="balance-info">
<div class="balance-row material-row">
<span>Material/Expense Total:</span>
<span>PHP {material_total:,.2f}</span>
</div>
<div class="balance-row" style="font-size: 13px; margin-top: 4px; border-top: 1px solid rgba(255,255,255,0.2); padding-top: 4px;">
<span>Excess Money Total:</span>
<span style="color: #a5d6a7;">PHP {excess_total:,.2f}</span>
</div>
<div class="balance-row" style="font-size: 13px;">
<span>Total Budget:</span>
<span>PHP {budget:,.2f}</span>
</div>
"""
    if sobra_amount > 0:
        html += f"""
<div class="final-balance-row">
<span>EXCESS</span>
<span style="color: #a5d6a7;">PHP {sobra_amount:,.2f}</span>
</div>"""
    if kulang_amount > 0:
        html += f"""
<div class="final-balance-row">
<span>SHORTAGE</span>
<span style="color: #e57373;">PHP {kulang_amount:,.2f}</span>
</div>"""

    html += f"""
<div class="final-balance-row">
<span>FINAL BALANCE</span>
<span style="color: {balance_color};">PHP {remaining_balance:,.2f}</span>
</div>
</div>
</div>
</div>
</div>
<div class="footer">
This document was electronically generated and is valid without signature.
</div>
</div>
<script>
function saveAsImage() {{
    const element = document.getElementById('receiptContent');
    html2canvas(element, {{ scale: 2, useCORS: true }}).then(canvas => {{
        const link = document.createElement('a');
        link.download = '{custom_title.replace(" ", "_")}_Receipt.png';
        link.href = canvas.toDataURL('image/png');
        link.click();
    }});
}}
</script>
</body>
</html>"""
    return html


def generate_payroll_html(labor_records, expense_records, remaining_money=0.0, custom_title="INVENTORY RECEIPT"):
    date_str = manila_now().strftime("%B %d, %Y | %I:%M %p")
    total_labor = sum(r['net'] for r in labor_records)
    total_expenses = sum(e['price'] for e in expense_records)
    sub_total = total_labor + total_expenses
    grand_total = sub_total - (remaining_money or 0.0)

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<style>
* {{ box-sizing: border-box; }}
html, body {{ width: 100%; max-width: 100%; margin: 0; overflow-x: hidden; }}
@import url('https://fonts.googleapis.com/css?family=Inter:wght@400;600;700&display=swap');
body {{ font-family: 'Inter', sans-serif !important; background-color: #f0f4f0 !important; color: #333; padding: 20px !important; }}
#receiptContent {{ width: min(100%, 1000px); margin: 0 auto !important; background: #fff !important; padding: 30px !important; border-radius: 4px !important; box-shadow: 0 4px 20px rgba(0,0,0,0.08) !important; border-top: 10px solid #1b5e20 !important; }}
#receiptContent table {{ max-width: 100%; }}
#receiptContent th {{ background-color: #1b5e20 !important; color: #fff !important; text-transform: uppercase; letter-spacing: 1px; }}
#receiptContent td {{ border-bottom: 1px solid #f0f0f0 !important; overflow-wrap: anywhere; }}
#receiptContent h1, #receiptContent h3 {{ color: #1b5e20 !important; }}
#receiptContent > table:first-child {{ margin-bottom: 30px !important; }}
#receiptContent > table:last-of-type td:last-child {{ background: #013220 !important; color: #fff !important; }}
.save-btn-container {{ text-align: center; margin-bottom: 25px; }}
.save-img-btn {{ background-color: #1b5e20; color: white; border: none; padding: 12px 24px; font-size: 14px; font-weight: bold; border-radius: 6px; cursor: pointer; box-shadow: 0 4px 10px rgba(0,0,0,0.15); }}
.save-img-btn:hover {{ background-color: #2e7d32; }}
@media print {{ .save-btn-container {{ display: none; }} }}
@media (max-width: 600px) {{
    body {{ padding: 10px !important; }}
    #receiptContent {{ width: 100%; padding: 16px !important; text-align: center; }}
    #receiptContent > table {{ table-layout: fixed; font-size: 10px; }}
    #receiptContent th, #receiptContent td {{ padding: 7px 3px !important; line-height: 1.25; }}
    #receiptContent h1 {{ font-size: 20px !important; }}
    #receiptContent h3 {{ font-size: 15px !important; }}
    #receiptContent > table:first-child td {{ display: block; width: 100% !important; text-align: center !important; }}
    #receiptContent > table:last-of-type td:last-child {{ width: 100% !important; text-align: center !important; }}
    .save-img-btn {{ max-width: 100%; padding: 10px 12px; font-size: 12px; }}
}}
</style>
</head>
<body style="font-family: 'Segoe UI', sans-serif; background-color: #f4f7f6; padding: 40px;">
<div class="save-btn-container">
<button class="save-img-btn" onclick="saveAsImage()">SEE PHOTO & DOWNLOAD IMAGE (Phone & Laptop)</button>
</div>
<div id="receiptContent" style="max-width: 900px; margin: auto; background: white; border-top: 10px solid #1b5e20; padding: 40px; border-radius: 8px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
<table style="width: 100%; border-collapse: collapse; margin-bottom: 30px;">
<tr>
<td>
<h1 style="color: #1b5e20; margin: 0; text-transform: uppercase;">Ailyn House Project</h1>
<p style="color: #555; margin: 5px 0 0 0;">Official Labor Tally</p>
<p style="color: #777; font-size: 14px; margin: 0;">Management System v3.6 Enterprise</p>
</td>
<td style="text-align: right;">
<h3 style="color: #1b5e20; margin: 0;">{custom_title}</h3>
<p style="color: #555; font-size: 14px; margin: 5px 0 0 0;">Date: {date_str}</p>
<p style="color: #777; font-size: 12px; margin: 5px 0 0 0;">Account: {RECEIVER_EMAIL}</p>
</td>
</tr>
</table>
<div style="border-bottom: 2px solid #eee; margin-bottom: 30px;"></div>
<table style="width: 100%; border-collapse: collapse; margin-bottom: 30px;">
<thead>
<tr style="background-color: #1b5e20; color: white; text-transform: uppercase; font-size: 14px;">
<th style="padding: 12px; text-align: left;">Worker Name</th>
<th style="padding: 12px; text-align: center;">Role</th>
<th style="padding: 12px; text-align: center;">Days / Point</th>
<th style="padding: 12px; text-align: right;">Gross Pay</th>
<th style="padding: 12px; text-align: right;">C.A.</th>
<th style="padding: 12px; text-align: right;">Net Pay</th>
</tr>
</thead>
<tbody>"""

    for r in labor_records:
        role_display = r.get('role', 'Labor')
        gross = r.get('gross_pay', r['days'] * r['rate'])
        html += f"""
<tr>
<td style="padding: 12px; border-bottom: 1px solid #ddd; font-weight: bold;">{r['name']}</td>
<td style="padding: 12px; border-bottom: 1px solid #ddd; text-align: center;">{role_display}</td>
<td style="padding: 12px; border-bottom: 1px solid #ddd; text-align: center;">{r['days']:.1f}</td>
<td style="padding: 12px; border-bottom: 1px solid #ddd; text-align: right;">{gross:,.2f}</td>
<td style="padding: 12px; border-bottom: 1px solid #ddd; text-align: right; color: #d32f2f;">({r['ca']:,.2f})</td>
<td style="padding: 12px; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold; color: #1b5e20;">{r['net']:,.2f}</td>
</tr>"""

    if expense_records:
        html += f"""
<tr>
<td colspan="6" style="padding: 12px 0;"></td>
</tr>
<tr style="background-color: #388e3c; color: white; text-transform: uppercase; font-size: 14px;">
<th colspan="5" style="padding: 10px; text-align: left;">Expense Description</th>
<th style="padding: 10px; text-align: right;">Amount</th>
</tr>"""
        for e in expense_records:
            html += f"""
<tr>
<td colspan="5" style="padding: 10px; border-bottom: 1px solid #ddd;">{e['item']}</td>
<td style="padding: 10px; border-bottom: 1px solid #ddd; text-align: right; font-weight: bold;">{e['price']:,.2f}</td>
</tr>"""

    html += f"""
</tbody>
</table>
<table style="width: 100%; border-collapse: collapse; margin-top: 20px; margin-bottom: 30px;">
<tr style="border-top: 2px solid #bbb;">
<td style="padding: 12px; font-weight: bold; text-align: right; font-size: 15px;">Subtotal Expenses:</td>
<td style="padding: 12px; width: 180px; text-align: right; font-weight: bold; font-size: 15px; color: #333;">PHP {sub_total:,.2f}</td>
</tr>"""

    if remaining_money and remaining_money > 0:
        html += f"""
<tr style="border-bottom: 2px solid #bbb;">
<td style="padding: 12px; font-weight: bold; text-align: right; color: #d32f2f; font-size: 15px;">Remaining/Leftover Money:</td>
<td style="padding: 12px; width: 180px; text-align: right; font-weight: bold; color: #d32f2f; font-size: 15px;">-PHP {remaining_money:,.2f}</td>
</tr>"""

    html += f"""
</table>
<table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
<tr>
<td></td>
<td style="width: 350px; background: #1b5e20; color: white; padding: 20px; border-radius: 8px; text-align: right;">
<span style="font-size: 14px; text-transform: uppercase; letter-spacing: 1px;">Final Output Amount</span><br>
<span style="font-size: 32px; font-weight: bold; margin-top: 5px; display: inline-block;">PHP {grand_total:,.2f}</span>
</td>
</tr>
</table>
<div style="text-align: center; margin-top: 60px; border-top: 1px solid #eee; padding-top: 20px;">
<p style="color: #999; font-size: 11px; letter-spacing: 1px; text-transform: uppercase;">
THIS DOCUMENT WAS ELECTRONICALLY GENERATED AND IS VALID WITHOUT SIGNATURE.
</p>
</div>
</div>
<script>
function saveAsImage() {{
    const element = document.getElementById('receiptContent');
    html2canvas(element, {{ scale: 2, useCORS: true }}).then(canvas => {{
        const link = document.createElement('a');
        link.download = '{custom_title.replace(" ", "_")}_Receipt.png';
        link.href = canvas.toDataURL('image/png');
        link.click();
    }});
}}
</script>
</body>
</html>"""
    return html, grand_total


def clear_all():
    st.session_state.records = []
    st.session_state.labor_records = []
    st.session_state.payroll_expenses = []
    st.session_state.planner_tasks = []
    st.session_state.budget = 0.0
    st.session_state.budget_history = []
    st.session_state.remaining_money = 0.0
    st.session_state.receipt_archive = []
    st.session_state.client_notes = []
    for photo in st.session_state.get("scanner_photos", []):
        delete_scanner_photo(photo.get("file", ""))
    st.session_state.scanner_photos = []
    for report_type in ("construction", "payroll"):
        for report_path in list_saved_reports(report_type):
            if os.path.exists(report_path):
                os.remove(report_path)
    st.session_state.view = "home"
    st.session_state.selected_role = "Labor"
    save_state(st.session_state)


def persist_state():
    save_state(st.session_state)


@st.dialog("Take Photo")
def photo_camera_dialog():
    st.markdown("""
    <style>
    div[role="dialog"] { background: #080b0c; border: 1px solid #33423d; border-radius: 24px; padding: 1rem 1rem .85rem; }
    div[role="dialog"] [data-testid="stVerticalBlock"] { gap: 0.55rem; }
    div[role="dialog"] h2 { color: #f4fff7; font-size: 1.35rem; letter-spacing: .01em; }
    div[role="dialog"] [data-testid="stRadio"] label p { font-size: 12px; font-weight: 700; }
    div[role="dialog"] [data-testid="stCaptionContainer"] p { color: #9fb0a6; font-size: 10px; }
    div[role="dialog"] [data-testid="stCameraInput"] { width: 100%; }
    div[role="dialog"] [data-testid="stCameraInput"] video,
    div[role="dialog"] [data-testid="stCameraInput"] img { width: 100% !important; max-height: 52vh; object-fit: cover; border-radius: 14px; }
    div[role="dialog"] button { min-height: 44px !important; font-size: 12px !important; }
    @media (max-width: 600px) {
        div[role="dialog"] { width: calc(100vw - 20px) !important; max-width: calc(100vw - 20px) !important; margin: 10px !important; padding: .8rem .7rem .7rem; }
        div[role="dialog"] h2 { font-size: 1.15rem; }
        div[role="dialog"] [data-testid="stRadio"] > div { gap: 4px !important; }
        div[role="dialog"] [data-testid="stRadio"] label { padding-right: 4px !important; }
        div[role="dialog"] [data-testid="stRadio"] label p { font-size: 11px; }
        div[role="dialog"] [data-testid="stCameraInput"] video,
        div[role="dialog"] [data-testid="stCameraInput"] img { max-height: 45vh; }
        div[role="dialog"] [data-testid="stDataFrame"] { max-width: 100%; overflow-x: auto; }
    }
    </style>
    """, unsafe_allow_html=True)
    flash_mode = st.radio(
        "FLASHLIGHT",
        ["Auto", "On", "Off"],
        horizontal=True,
        index=["Auto", "On", "Off"].index(st.session_state.scanner_flash_mode),
        key="camera_flash_mode",
    )
    st.session_state.scanner_flash_mode = flash_mode
    st.caption(f"Flash: {flash_mode.upper()}")
    camera_mode = st.radio(
        "CAMERA",
        ["Back camera", "Front camera"],
        horizontal=True,
        index=["Back camera", "Front camera"].index(st.session_state.scanner_camera_mode),
        key="camera_lens_mode",
    )
    st.session_state.scanner_camera_mode = camera_mode
    st.caption(f"{camera_mode} selected. If your phone opens the other lens, tap the camera switch icon.")
    photo_category = st.selectbox(
        "WORK CATEGORY",
        ["General", "Before", "After", "Framing", "Electrical", "Plumbing", "Painting", "Inspection"],
        key="camera_photo_category",
    )
    photo = st.camera_input(f"Take project photo with {camera_mode.lower()}", key=f"modal_photo_scanner_{st.session_state.scanner_input_version}")
    if photo:
        photo_hash = hashlib.sha256(photo.getvalue()).hexdigest()
        if st.session_state.get("scanned_photo_hash") != photo_hash:
            st.session_state.scanned_photo_bytes = photo.getvalue()
            st.session_state.scanned_photo_mime = photo.type or "image/jpeg"
            try:
                st.session_state.scanned_photo_text = scan_photo_text(photo)
            except TESSERACT_NOT_FOUND_ERROR:
                st.session_state.scanned_photo_text = "OCR is unavailable. Install Tesseract OCR on the server."
            except RuntimeError as error:
                st.session_state.scanned_photo_text = str(error)
            except (OSError, ValueError):
                st.session_state.scanned_photo_text = "The photo could not be read. Try taking it again."
            st.session_state.scanned_photo_hash = photo_hash

    if st.session_state.get("scanned_photo_bytes"):
        view_col, delete_col = st.columns(2)
        with view_col:
            if st.button("◀ VIEW PHOTO", use_container_width=True, key="modal_view_photo"):
                st.session_state.show_scanned_photo = True
        with delete_col:
            if st.button("DELETE ▶", use_container_width=True, key="modal_delete_photo"):
                photo_hash = st.session_state.get("scanned_photo_hash")
                kept_photos = []
                for saved_photo in st.session_state.scanner_photos:
                    if saved_photo.get("hash") == photo_hash:
                        delete_scanner_photo(saved_photo.get("file", ""))
                    else:
                        kept_photos.append(saved_photo)
                st.session_state.scanner_photos = kept_photos
                persist_state()
                st.session_state.scanned_photo_bytes = None
                st.session_state.scanned_photo_hash = None
                st.session_state.scanned_photo_text = ""
                st.session_state.scanner_input_version += 1
                st.rerun()
        if st.session_state.get("show_scanned_photo"):
            st.image(st.session_state.scanned_photo_bytes, caption="Captured photo", use_container_width=True)
        save_col, retake_col = st.columns(2)
        with save_col:
            if st.button("SAVE FILE", use_container_width=True, key="modal_save_photo"):
                photo_hash = st.session_state.get("scanned_photo_hash")
                if not any(photo.get("hash") == photo_hash for photo in st.session_state.scanner_photos):
                    photo_id = str(uuid.uuid4())
                    relative_path = save_scanner_photo(st.session_state.scanned_photo_bytes, st.session_state.get("scanned_photo_mime", "image/jpeg"), photo_id)
                    st.session_state.scanner_photos.append({"id": photo_id, "hash": photo_hash, "file": relative_path, "tag": photo_category, "saved_at": manila_now().isoformat()})
                    persist_state()
                    st.success("Photo saved to the project archive.")
                else:
                    st.info("This photo is already saved.")
        with retake_col:
            if st.button("RETAKE", use_container_width=True, key="modal_retake_photo"):
                st.session_state.scanned_photo_bytes = None
                st.session_state.scanned_photo_hash = None
                st.session_state.scanned_photo_text = ""
                st.session_state.show_scanned_photo = False
                st.session_state.scanner_input_version += 1
                st.rerun()
        scanned_text = st.session_state.get("scanned_photo_text", "")
        if scanned_text:
            if scanned_text.startswith("OCR is unavailable") or scanned_text.startswith("The photo could not be read"):
                st.warning(scanned_text)
            else:
                scanned_fields = parse_scanned_receipt(scanned_text)
                st.success("PHOTO SCANNED AND ENCODED")
                st.markdown("#### DETECTED RECEIPT DETAILS")
                st.table({
                    "Field": ["Item", "Quantity", "Unit price", "Delivery", "Total"],
                    "Encoded value": [
                        scanned_fields["name"] or "Not detected",
                        scanned_fields["qty"],
                        f"PHP {scanned_fields['price']:,.2f}",
                        f"PHP {scanned_fields['delivery']:,.2f}",
                        f"PHP {(scanned_fields['price'] * scanned_fields['qty'] + scanned_fields['delivery']):,.2f}",
                    ],
                })
            with st.expander("View scanned text", expanded=False):
                st.text_area("Recognized text", value=scanned_text, height=120, key="modal_scanned_text", disabled=True, label_visibility="collapsed")
            if not scanned_text.startswith("OCR is unavailable") and not scanned_text.startswith("The photo could not be read") and st.button("USE SCAN IN MATERIAL ENTRY", use_container_width=True, key="modal_use_scanned_material"):
                st.session_state.material_name = scanned_fields["name"]
                st.session_state.material_price = scanned_fields["price"] or None
                st.session_state.material_qty = scanned_fields["qty"]
                st.session_state.material_delivery = scanned_fields["delivery"] or None
                set_view("material")
    if st.button("CLOSE CAMERA", use_container_width=True, key="close_photo_dialog"):
        st.session_state.scanner_open = False
        st.session_state.show_scanned_photo = False
        st.rerun()


def install_update(uploaded_file, signature):
    """Validate and atomically install an uploaded app upgrade after a backup."""
    source = uploaded_file.getvalue()
    if not source:
        raise ValueError("The uploaded upgrade file is empty.")
    try:
        ast.parse(source.decode("utf-8"), filename=uploaded_file.name)
    except (UnicodeDecodeError, SyntaxError) as error:
        raise ValueError(f"The upgrade was rejected: {error}") from None
    if not UPDATE_SIGNING_KEY:
        raise ValueError("Updates are disabled until AILYN_UPDATE_SIGNING_KEY is configured.")
    expected_signature = hmac.new(
        UPDATE_SIGNING_KEY.encode("utf-8"), source, hashlib.sha256
    ).hexdigest()
    if not hmac.compare_digest(expected_signature, signature.strip()):
        raise ValueError("The upgrade signature is invalid.")

    backup_dir = os.path.join(APP_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    backup_path = os.path.join(backup_dir, f"NELL.py.py.{int(time.time())}.bak")
    temporary_path = None
    try:
        shutil.copy2(__file__, backup_path)
        with tempfile.NamedTemporaryFile("wb", delete=False, dir=APP_DIR,
                                         prefix=".nell-upgrade-") as temporary:
            temporary.write(source)
            temporary_path = temporary.name
        os.replace(temporary_path, __file__)
    except OSError as error:
        if temporary_path and os.path.exists(temporary_path):
            os.remove(temporary_path)
        raise ValueError(f"The upgrade could not be installed: {error}") from None
    return backup_path


def record_budget_change(action, amount, previous_budget):
    st.session_state.budget_history.append({
        "date": manila_now().strftime("%b %d, %Y %I:%M %p"),
        "action": action,
        "amount": float(amount),
        "previous": float(previous_budget),
        "total": float(st.session_state.budget),
    })
    persist_state()


CALCULATOR_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
    ast.Pow: operator.pow,
    ast.USub: operator.neg,
    ast.UAdd: operator.pos,
}


def calculate_expression(expression):
    def evaluate(node):
        if isinstance(node, ast.Expression):
            return evaluate(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.BinOp) and type(node.op) in CALCULATOR_OPERATORS:
            left = evaluate(node.left)
            right = evaluate(node.right)
            if isinstance(node.op, ast.Div) and right == 0:
                raise ValueError("Cannot divide by zero.")
            if isinstance(node.op, ast.Pow) and abs(right) > 10:
                raise ValueError("Exponent is limited to 10.")
            return CALCULATOR_OPERATORS[type(node.op)](left, right)
        if isinstance(node, ast.UnaryOp) and type(node.op) in CALCULATOR_OPERATORS:
            return CALCULATOR_OPERATORS[type(node.op)](evaluate(node.operand))
        raise ValueError("Use numbers and +, -, *, /, or ** only.")

    try:
        result = evaluate(ast.parse(expression, mode="eval"))
    except (SyntaxError, ValueError, TypeError, ZeroDivisionError, OverflowError):
        raise ValueError("Enter a valid arithmetic expression.") from None
    if not result or abs(result) <= 1e12:
        return result
    raise ValueError("Result is limited to 1,000,000,000,000.")


@st.dialog("Budget Control", width="small")
def budget_dialog():
    st.caption(f"Current budget: PHP {float(st.session_state.budget):,.2f}")
    apply_tab, edit_tab, calculator_tab, history_tab = st.tabs(
        ["Apply", "Edit", "Calculator", "History"]
    )

    with apply_tab:
        with st.form("budget_apply_dialog_form", clear_on_submit=True):
            amount = st.number_input("Amount to add", min_value=0.01, value=None,
                                     placeholder="0.00")
            submitted = st.form_submit_button("Apply Budget", use_container_width=True)
        if submitted:
            if amount is not None:
                previous_budget = float(st.session_state.budget)
                st.session_state.budget = previous_budget + float(amount)
                record_budget_change("Applied", amount, previous_budget)
                st.success("Budget applied.")
                st.rerun()
            else:
                st.warning("Enter an amount first.")

    with edit_tab:
        with st.form("budget_edit_dialog_form"):
            edited_budget = st.number_input(
                "New total budget", min_value=0.0,
                value=float(st.session_state.budget), step=100.0,
            )
            edited = st.form_submit_button("Save Budget Edit", use_container_width=True)
        if edited:
            previous_budget = float(st.session_state.budget)
            st.session_state.budget = float(edited_budget)
            record_budget_change("Edited", edited_budget - previous_budget, previous_budget)
            st.success("Budget updated.")
            st.rerun()

    with calculator_tab:
        st.caption("Use expressions such as 12500 - 3200 + 450 or 1500 * 2.")
        if st.button("Clear", key="budget_calculator_clear", use_container_width=True):
            st.session_state["budget_calculator_expression"] = ""
            st.rerun()
        with st.form("budget_calculator_form"):
            expression = st.text_input(
                "Expression", placeholder="0 + 0", key="budget_calculator_expression"
            )
            calculate = st.form_submit_button("Calculate", use_container_width=True)
        if calculate:
            try:
                result = calculate_expression(expression)
                st.metric("Result", f"PHP {result:,.2f}")
            except ValueError as error:
                st.warning(str(error))

    with history_tab:
        history = st.session_state.budget_history
        if history:
            st.dataframe(
                [{"Date": entry["date"], "Action": entry["action"],
                  "Change": f"PHP {entry['amount']:,.2f}",
                  "Total": f"PHP {entry['total']:,.2f}"} for entry in reversed(history)],
                use_container_width=True, hide_index=True,
            )
        else:
            st.info("No budget changes yet.")


@st.dialog("Add Labor Account", width="small")
def payroll_labor_dialog():
    role = st.selectbox("Role", list(FULL_DAY_RATES), key="popup_labor_role")
    worker = st.text_input("Worker name", key="popup_labor_worker")
    days = st.number_input("Worked days", min_value=0.0, value=None, placeholder="Enter days", key="popup_labor_days")
    cash_advance = st.number_input("Cash advance", min_value=0.0, value=0.0, step=50.0, key="popup_labor_ca")
    if st.button("SAVE LABOR ACCOUNT", use_container_width=True, key="popup_save_labor"):
        worked_days = float(days or 0.0)
        if not worker.strip() or worked_days <= 0:
            st.warning("Enter a worker name and valid worked days.")
            return
        gross_pay, full_pay, partial_pay = calculate_labor_pay(worked_days, role)
        st.session_state.labor_records.append({
            "id": str(uuid.uuid4()),
            "date": manila_now().strftime("%b %d, %Y"),
            "recorded_at": manila_now().isoformat(),
            "name": worker.strip().upper(),
            "role": role,
            "days": worked_days,
            "rate": FULL_DAY_RATES[role],
            "gross_pay": gross_pay,
            "full_pay": full_pay,
            "partial_pay": partial_pay,
            "ca": float(cash_advance),
            "net": gross_pay - float(cash_advance),
        })
        persist_state()
        st.success(f"{worker.strip().upper()} saved. Net pay: PHP {gross_pay - float(cash_advance):,.2f}")


@st.dialog("Payroll Ledger", width="medium")
def payroll_ledger_dialog():
    if not st.session_state.labor_records and not st.session_state.payroll_expenses:
        st.info("No payroll records yet.")
        return
    rows = [
        {
            "Date": record.get("date", ""),
            "Worker": record.get("name", ""),
            "Role": record.get("role", ""),
            "Days": f"{float(record.get('days', 0)):.1f}",
            "Net Pay": f"PHP {float(record.get('net', 0)):,.2f}",
        }
        for record in reversed(st.session_state.labor_records)
    ]
    rows.extend({
        "Date": record.get("date", ""),
        "Worker": record.get("item", ""),
        "Role": "Payroll Expense",
        "Days": "-",
        "Net Pay": f"PHP {float(record.get('price', 0)):,.2f}",
    } for record in reversed(st.session_state.payroll_expenses))
    st.dataframe(rows, use_container_width=True, hide_index=True)
    if st.button("OPEN FULL PAYROLL LEDGER", use_container_width=True, key="popup_open_payroll_ledger"):
        set_view("payroll_ledger")


@st.dialog("Payroll Report Summary", width="small")
def payroll_report_dialog():
    labor_total = sum(float(record.get("net", 0)) for record in st.session_state.labor_records)
    expense_total = sum(float(record.get("price", 0)) for record in st.session_state.payroll_expenses)
    remainder = float(st.session_state.remaining_money or 0)
    st.dataframe([
        {"Item": "Workers", "Value": len(st.session_state.labor_records)},
        {"Item": "Net labor", "Value": f"PHP {labor_total:,.2f}"},
        {"Item": "Payroll expenses", "Value": f"PHP {expense_total:,.2f}"},
        {"Item": "Remainder", "Value": f"PHP {remainder:,.2f}"},
        {"Item": "Final payout", "Value": f"PHP {labor_total + expense_total - remainder:,.2f}"},
    ], use_container_width=True, hide_index=True)
    if st.button("OPEN FULL PAYROLL REPORT", use_container_width=True, key="popup_open_payroll_report"):
        set_view("payroll_export")


def add_tx(name, price, qty, delivery, ttype, sender, record_date=None, details=None):
    p = float(price or 0.0)
    q = int(qty or 0)
    d = float(delivery or 0.0)
    if p <= 0 or q <= 0:
        return False
    amount = (p * q) + d if ttype == "material" else p
    st.session_state.records.append({
        "id": str(time.time()),
        "date": manila_now().strftime("%b %d, %Y"),
        "recorded_at": datetime.combine(record_date or manila_now().date(), datetime.min.time(), PHILIPPINES_TZ).isoformat(),
        "name": name.upper(),
        "price": p,
        "qty": q,
        "delivery": d,
        "amount": float(amount),
        "type": ttype,
        "sender": sender,
        **(details or {}),
    })
    persist_state()
    return True


st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Outfit:wght@500;600;700;800;900&display=swap');
:root{
  --mint:#72f7b0; --mint2:#c8ffe0; --deep:#06170f;
  --panel:rgba(6,27,18,.68); --panel2:rgba(10,42,26,.76);
  --edge:rgba(191,255,216,.20); --white:#f5fff8; --muted:#a9c9b5;
  --green:#0b6b2d; --glass:rgba(6,27,18,.58);
}
*{box-sizing:border-box}
html,body,[class*="css"]{font-family:'Manrope',sans-serif}
.stApp{
  background:url("https://images.unsplash.com/photo-1600585154340-be6161a56a0c") no-repeat center center fixed;
  background-size:cover;background-position:center;
}
.stApp:before{
  content:"";position:fixed;inset:0;z-index:0;pointer-events:none;
  background:
    radial-gradient(circle at 12% 10%,rgba(74,222,128,.10),transparent 30%),
    radial-gradient(circle at 88% 78%,rgba(20,184,166,.10),transparent 32%),
    linear-gradient(115deg,rgba(1,8,5,.38),rgba(3,25,15,.58));
}
.stApp:after{
  content:"";position:fixed;inset:0;z-index:0;pointer-events:none;opacity:.055;
  background-image:linear-gradient(rgba(255,255,255,.4) 1px,transparent 1px),linear-gradient(90deg,rgba(255,255,255,.4) 1px,transparent 1px);
  background-size:46px 46px;mask-image:linear-gradient(to bottom,black,transparent 82%);
}
.block-container{
  position:relative;z-index:1;max-width:1500px!important;
  padding:28px 34px 42px!important;margin:18px auto 28px!important;
  background:rgba(3,18,11,.58)!important;
  border:1px solid rgba(210,255,225,.13);border-radius:34px;
  backdrop-filter:blur(30px) saturate(145%);-webkit-backdrop-filter:blur(30px) saturate(145%);
  box-shadow:0 35px 90px rgba(0,0,0,.48),inset 0 1px 0 rgba(255,255,255,.10);
}
/* premium glass header */
.headbar-container{display:flex;justify-content:center;margin:0 auto 28px}
.headbar-card{
  width:100%;position:relative;overflow:hidden;padding:20px 24px 18px;
  display:flex;align-items:center;justify-content:space-between;gap:20px;
  border-radius:26px;background:linear-gradient(105deg,rgba(10,55,32,.90),rgba(4,25,16,.78));
  border:1px solid rgba(166,255,197,.25);
  box-shadow:0 18px 42px rgba(0,0,0,.38),inset 0 1px 0 rgba(255,255,255,.15);
  backdrop-filter:blur(18px) saturate(135%);-webkit-backdrop-filter:blur(18px) saturate(135%);
}
.headbar-card:before{content:"";position:absolute;inset:0;background:linear-gradient(110deg,transparent 0%,rgba(255,255,255,.08) 42%,transparent 54%);transform:translateX(-120%);animation:scan 8s linear infinite}
.headbar-card:after{content:"";position:absolute;right:-90px;top:-100px;width:260px;height:260px;border-radius:50%;background:radial-gradient(circle,rgba(114,247,176,.13),transparent 68%);pointer-events:none}
.headbar-title{position:relative;z-index:1;display:flex;align-items:center;gap:16px;color:#fff!important;font-family:'Outfit';font-size:28px!important;font-weight:900;letter-spacing:.055em;line-height:1}
.headbar-title img{width:76px;height:76px;object-fit:contain;filter:drop-shadow(0 8px 16px rgba(0,0,0,.28));transition:transform .25s ease,filter .25s ease}
.headbar-title:hover img{transform:translateY(-3px) scale(1.04);filter:drop-shadow(0 12px 20px rgba(114,247,176,.28))}
.headbar-subtitle{position:relative;z-index:1;color:#aeeec3;font-size:10px;letter-spacing:.16em;text-transform:uppercase;font-weight:800;margin-left:92px;margin-top:4px}
.headbar-time{position:relative;z-index:1;color:#c8ffe0!important;font-size:11px!important;font-weight:800!important;background:rgba(255,255,255,.06);border:1px solid rgba(173,255,201,.15);padding:10px 13px;border-radius:13px;box-shadow:inset 0 1px 0 rgba(255,255,255,.08)}
/* glass sidebar */
section[data-testid="stSidebar"]{background:linear-gradient(180deg,rgba(1,13,8,.92),rgba(3,24,14,.88))!important;border-right:1px solid rgba(114,247,176,.12)!important;box-shadow:18px 0 70px rgba(0,0,0,.52)!important;backdrop-filter:blur(22px) saturate(140%)!important;-webkit-backdrop-filter:blur(22px) saturate(140%)!important}
section[data-testid="stSidebar"]>div{padding:22px 14px 30px!important} section[data-testid="stSidebar"] *{color:#edfff3!important}
.sidebar-brand{position:relative;overflow:hidden;padding:18px 16px;border-radius:26px;margin-bottom:12px;background:linear-gradient(145deg,rgba(12,65,39,.54),rgba(2,23,14,.38));border:1px solid rgba(114,247,176,.20);box-shadow:0 20px 44px rgba(0,0,0,.38),inset 0 1px 0 rgba(255,255,255,.12),0 0 32px rgba(52,211,125,.06);backdrop-filter:blur(22px) saturate(145%);-webkit-backdrop-filter:blur(22px) saturate(145%);transition:.25s cubic-bezier(.2,.8,.2,1)} .sidebar-brand:hover{transform:translateY(-3px);border-color:rgba(114,247,176,.40);box-shadow:0 26px 52px rgba(0,0,0,.44),0 0 34px rgba(114,247,176,.12),inset 0 1px 0 rgba(255,255,255,.16)}
.sidebar-brand:after{content:"";position:absolute;inset:-80% 35%;background:rgba(255,255,255,.09);transform:rotate(25deg);animation:scan 7s linear infinite}
.brand-row{position:relative;z-index:1;display:flex;align-items:center;gap:14px}.brand-logo{width:62px;height:62px;object-fit:contain;filter:drop-shadow(0 8px 16px rgba(0,0,0,.32));transition:.25s ease}.sidebar-brand:hover .brand-logo{transform:translateY(-4px) scale(1.06);filter:drop-shadow(0 14px 26px rgba(114,247,176,.30))}.brand-copy{min-width:0}.brand-title{font-family:'Outfit';font-size:17px;font-weight:900;letter-spacing:.06em;line-height:1.02;color:#fff!important}.brand-title span{display:block}.brand-sub{font-size:9px;color:#8ff1b4!important;letter-spacing:.16em;text-transform:uppercase;margin-top:7px;font-weight:800}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3{font-family:'Outfit';font-size:10px!important;letter-spacing:.18em;text-transform:uppercase;color:#72f7b0!important;margin:20px 5px 9px!important;display:flex;align-items:center;gap:9px} section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3:before{content:'•';font-size:20px;line-height:0;color:#45f39a;text-shadow:0 0 12px rgba(69,243,154,.8)} section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3:after{content:'';height:1px;flex:1;background:linear-gradient(90deg,rgba(114,247,176,.35),transparent)}section[data-testid="stSidebar"] hr{border-color:rgba(170,255,198,.10)!important;margin:10px 4px!important}.sidebar-gap{height:8px}.sidebar-live{font-size:9px;letter-spacing:.14em;color:#7feeb0!important;font-weight:800;text-align:center;margin:-4px 0 10px;text-shadow:0 0 12px rgba(114,247,176,.18)}
section[data-testid="stSidebar"] button{min-height:52px!important;margin:7px 0!important;padding:0 16px!important;border-radius:18px!important;text-align:left!important;background:linear-gradient(145deg,rgba(18,82,48,.48),rgba(2,31,18,.42))!important;border:1px solid rgba(114,247,176,.15)!important;box-shadow:0 7px 0 rgba(1,12,7,.55),0 14px 28px rgba(0,0,0,.24),inset 0 1px 0 rgba(255,255,255,.10)!important;transition:transform .18s cubic-bezier(.2,.8,.2,1),box-shadow .18s,border-color .18s,background .18s!important;backdrop-filter:blur(14px) saturate(135%)!important;-webkit-backdrop-filter:blur(14px) saturate(135%)!important}
section[data-testid="stSidebar"] button:hover{transform:translate3d(5px,-3px,0)!important;background:linear-gradient(145deg,rgba(28,116,67,.72),rgba(5,47,27,.56))!important;border-color:rgba(114,247,176,.58)!important;box-shadow:0 10px 0 rgba(1,12,7,.70),0 20px 36px rgba(0,0,0,.34),0 0 28px rgba(70,230,132,.16),inset 0 1px 0 rgba(255,255,255,.18)!important}
section[data-testid="stSidebar"] button:active{transform:translate3d(2px,4px,0)!important;box-shadow:0 2px 0 rgba(1,12,7,.9),0 6px 12px rgba(0,0,0,.3)!important}
section[data-testid="stSidebar"] button p{font-family:'Manrope'!important;font-weight:800!important;font-size:12px!important;letter-spacing:.01em}
/* glass controls */
button,.stDownloadButton>button,.stFormSubmitButton>button{position:relative!important;overflow:hidden!important;min-height:46px!important;border-radius:16px!important;color:#f5fff8!important;font-weight:800!important;background:linear-gradient(145deg,rgba(25,92,54,.88),rgba(5,33,19,.94))!important;border:1px solid rgba(173,255,201,.22)!important;box-shadow:0 6px 0 rgba(2,17,10,.78),0 13px 27px rgba(0,0,0,.25),inset 0 1px 0 rgba(255,255,255,.13)!important;transition:all .17s ease!important}
button:hover,.stDownloadButton>button:hover,.stFormSubmitButton>button:hover{transform:translateY(-3px)!important;border-color:rgba(114,247,176,.58)!important;box-shadow:0 9px 0 rgba(2,17,10,.78),0 20px 34px rgba(0,0,0,.35),0 0 25px rgba(114,247,176,.13),inset 0 1px 0 rgba(255,255,255,.2)!important}
button:active,.stDownloadButton>button:active,.stFormSubmitButton>button:active{transform:translateY(3px)!important;box-shadow:0 2px 0 rgba(2,17,10,.8),0 5px 10px rgba(0,0,0,.28)!important}
/* glass input fields */
div[data-baseweb="input"],div[data-baseweb="base-input"],textarea,div[data-baseweb="select"]>div{background:rgba(3,25,15,.78)!important;border:1px solid rgba(163,255,194,.18)!important;border-radius:15px!important;color:#fff!important;min-height:48px!important;box-shadow:inset 0 4px 15px rgba(0,0,0,.23),0 4px 12px rgba(0,0,0,.12)!important}
input,textarea{color:#fff!important;-webkit-text-fill-color:#fff!important}input:focus,textarea:focus,div[data-baseweb="input"]:focus-within,div[data-baseweb="select"]>div:focus-within{border-color:#72f7b0!important;box-shadow:0 0 0 3px rgba(114,247,176,.08),0 0 24px rgba(114,247,176,.14),inset 0 4px 15px rgba(0,0,0,.2)!important}label{font-weight:700!important;color:#c8e8d2!important}
/* executive glass metrics */
[data-testid="stMetric"]{position:relative;overflow:hidden;min-height:118px;padding:20px!important;background:linear-gradient(145deg,rgba(17,76,44,.78),rgba(4,28,17,.82))!important;border:1px solid rgba(173,255,201,.20);border-radius:24px;box-shadow:0 16px 34px rgba(0,0,0,.30),inset 0 1px 0 rgba(255,255,255,.13);transition:.2s ease;backdrop-filter:blur(14px)}
[data-testid="stMetric"]:before,.dash-section:before,.dash-hero:before{content:"";position:absolute;inset:0;background:linear-gradient(110deg,transparent 0%,rgba(255,255,255,.06) 42%,transparent 54%);transform:translateX(-130%);animation:scan 9s linear infinite;pointer-events:none}
[data-testid="stMetric"]:hover{transform:translateY(-5px) scale(1.01);box-shadow:0 23px 44px rgba(0,0,0,.38),0 0 28px rgba(114,247,176,.11)}
[data-testid="stMetric"] label{font-size:10px!important;letter-spacing:.14em;text-transform:uppercase;color:#91d9ab!important}[data-testid="stMetricValue"]{font-family:'Outfit';font-size:30px!important;font-weight:900!important;color:#fff!important}
/* dashboard hero */
.dash-hero{position:relative;overflow:hidden;background:linear-gradient(145deg,rgba(16,69,40,.80),rgba(4,27,16,.72));border:1px solid rgba(165,255,195,.18);border-radius:28px;padding:25px 28px;margin-bottom:18px;box-shadow:0 18px 42px rgba(0,0,0,.32),inset 0 1px 0 rgba(255,255,255,.10);backdrop-filter:blur(18px) saturate(135%);-webkit-backdrop-filter:blur(18px) saturate(135%)}
.dash-hero:after{content:"";position:absolute;right:-70px;top:-100px;width:280px;height:280px;border-radius:50%;background:radial-gradient(circle,rgba(114,247,176,.12),transparent 68%);pointer-events:none}.hero-row{position:relative;z-index:1;display:flex;align-items:center;gap:20px}.hero-logo{width:82px;height:82px;object-fit:contain;filter:drop-shadow(0 9px 17px rgba(0,0,0,.28));transition:.25s ease}.dash-hero:hover .hero-logo{transform:translateY(-3px) scale(1.04);filter:drop-shadow(0 13px 23px rgba(114,247,176,.24))}.hero-title{font-family:'Outfit';font-size:42px;line-height:.95;font-weight:900;color:#fff;letter-spacing:.015em}.hero-sub{font-size:11px;letter-spacing:.24em;color:#aeeec3;text-transform:uppercase;margin-top:9px;font-weight:800}.hero-rule{height:1px;background:rgba(191,255,216,.18);margin:18px 0}.welcome{font-size:14px;color:#c8e8d2}.welcome b{color:#72f7b0}
/* glass content cards */
.dash-section{position:relative;overflow:hidden;background:linear-gradient(145deg,rgba(10,42,26,.76),rgba(4,25,16,.68));border:1px solid rgba(165,255,195,.18);border-radius:24px;padding:20px 22px;box-shadow:0 16px 34px rgba(0,0,0,.28),inset 0 1px 0 rgba(255,255,255,.08);height:100%;backdrop-filter:blur(18px) saturate(130%);-webkit-backdrop-filter:blur(18px) saturate(130%);transition:.22s ease}.dash-section:hover{transform:translateY(-4px);border-color:rgba(114,247,176,.38);box-shadow:0 24px 46px rgba(0,0,0,.38),0 0 26px rgba(114,247,176,.08),inset 0 1px 0 rgba(255,255,255,.12)}
.section-title{font-family:'Outfit';font-size:17px;font-weight:900;color:#fff;letter-spacing:.02em;margin-bottom:16px}.section-head{display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid rgba(191,255,216,.14);padding-bottom:12px;margin-bottom:12px}.section-head span{color:#91d9ab!important}
.tx-row{display:flex;align-items:center;justify-content:space-between;padding:13px 0;border-bottom:1px solid rgba(191,255,216,.10);transition:.18s ease}.tx-row:hover{padding-left:7px;background:linear-gradient(90deg,rgba(114,247,176,.05),transparent);border-radius:10px}.tx-left{display:flex;gap:12px;align-items:center}.tx-icon{width:38px;height:38px;border-radius:50%;display:grid;place-items:center;background:rgba(114,247,176,.10);border:1px solid rgba(114,247,176,.20);color:#72f7b0;font-weight:900;box-shadow:inset 0 1px 0 rgba(255,255,255,.12)}.tx-name{font-weight:800;font-size:13px;color:#f5fff8}.tx-type{font-size:10px;color:#91b8a0;margin-top:3px}.tx-right{text-align:right;font-weight:900;color:#f5fff8}.tx-date{font-size:10px;color:#91b8a0;font-weight:500;margin-top:3px}
.donut-wrap{display:flex;align-items:center;gap:25px}.donut{width:190px;height:190px;border-radius:50%;background:conic-gradient(#72f7b0 0deg var(--p1),#55b978 var(--p1) var(--p2),#d8b64c var(--p2) var(--p3),#e46f5c var(--p3) 360deg);position:relative;flex:0 0 190px;box-shadow:0 14px 32px rgba(0,0,0,.28),inset 0 2px 4px rgba(255,255,255,.15)}.donut:before{content:"";position:absolute;inset:0;border-radius:50%;box-shadow:inset 0 0 0 8px rgba(255,255,255,.035),inset 0 -8px 18px rgba(0,0,0,.20)}.donut:after{content:"";position:absolute;inset:47px;background:rgba(4,27,16,.92);border:1px solid rgba(191,255,216,.14);border-radius:50%;box-shadow:inset 0 4px 15px rgba(0,0,0,.28),0 2px 10px rgba(0,0,0,.20)}.donut-center{position:absolute;z-index:2;inset:0;display:grid;place-content:center;text-align:center;font-family:'Outfit';font-weight:900;color:#fff}.donut-center small{font:500 11px Manrope;color:#91b8a0}.legend{flex:1}.legend-row{display:flex;justify-content:space-between;gap:12px;padding:8px 0;font-size:12px;color:#d7eee0}.legend-row b{color:#fff}.dot{width:10px;height:10px;border-radius:50%;display:inline-block;margin-right:8px;box-shadow:0 0 9px rgba(114,247,176,.25)}
.schedule{display:flex;align-items:center;gap:20px}.schedule-icon{width:64px;height:64px;border-radius:18px;background:rgba(114,247,176,.10);border:1px solid rgba(114,247,176,.22);display:grid;place-items:center;color:#72f7b0;font-size:28px;box-shadow:inset 0 1px 0 rgba(255,255,255,.10),0 9px 20px rgba(0,0,0,.20)}.schedule-title{font-family:'Outfit';font-size:17px;color:#fff;font-weight:900}.schedule-muted{font-size:12px;color:#91b8a0;margin-top:4px}.open-planner{margin-left:auto;background:linear-gradient(145deg,rgba(25,92,54,.88),rgba(5,33,19,.94));color:#fff;padding:12px 20px;border-radius:14px;font-weight:800;border:1px solid rgba(173,255,201,.22);box-shadow:0 6px 0 rgba(2,17,10,.65),0 12px 24px rgba(0,0,0,.25);transition:.18s ease}.open-planner:hover{transform:translateY(-3px);box-shadow:0 9px 0 rgba(2,17,10,.65),0 18px 30px rgba(0,0,0,.34);border-color:rgba(114,247,176,.55)}
/* planner cards keep the same glass language */
.cal-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:18px;margin-top:18px}.cal-card{position:relative;overflow:hidden;background:linear-gradient(145deg,rgba(16,69,40,.80),rgba(4,27,16,.85));border:1px solid rgba(165,255,195,.18);border-radius:24px;padding:20px;box-shadow:0 14px 34px rgba(0,0,0,.28),inset 0 1px 0 rgba(255,255,255,.08);transition:.22s cubic-bezier(.2,.8,.2,1);backdrop-filter:blur(14px)}.cal-card:hover{transform:translateY(-7px) scale(1.012);border-color:rgba(114,247,176,.45);box-shadow:0 24px 50px rgba(0,0,0,.40),0 0 30px rgba(114,247,176,.11)}.cal-date-badge{background:rgba(114,247,176,.10);color:#72f7b0;border:1px solid rgba(114,247,176,.28);padding:5px 11px;border-radius:999px;font-size:10px;font-weight:900}.cal-task-title{color:#fff;font-family:'Outfit';font-size:17px;font-weight:800}.cal-phase{color:#a8dcb8;font-size:12px}.cal-status-tag{font-size:9px;font-weight:900;padding:5px 10px;border-radius:999px;text-transform:uppercase}.badge-notstarted{background:rgba(255,255,255,.07);color:#d1d5db}.badge-inprogress{background:rgba(245,158,11,.14);color:#fbbf24}.badge-completed{background:rgba(34,197,94,.14);color:#65f394}
[data-testid="stExpander"]{background:rgba(5,29,17,.58)!important;border:1px solid rgba(163,255,194,.15)!important;border-radius:20px!important;box-shadow:0 10px 25px rgba(0,0,0,.18)!important}.stAlert{border-radius:17px!important;background:rgba(8,42,24,.65)!important;border:1px solid rgba(163,255,194,.18)!important}
@keyframes scan{0%,55%{transform:translateX(-130%)}80%,100%{transform:translateX(180%)}}
@media(max-width:900px){.block-container{padding:18px 14px 30px!important;margin:10px!important}.headbar-card{padding:16px}.headbar-title{font-size:24px!important}.headbar-subtitle{margin-left:92px}.hero-title{font-size:32px}.donut-wrap{flex-direction:column;align-items:flex-start}.donut{width:170px;height:170px;flex-basis:170px}.donut:after{inset:42px}.schedule{align-items:flex-start;flex-wrap:wrap}.open-planner{margin-left:0}}
@media(max-width:600px){.headbar-card{display:block}.headbar-title{font-size:21px!important;gap:10px}.headbar-title img{width:52px;height:52px}.headbar-subtitle{margin-left:62px;font-size:8px}.headbar-time{margin-top:12px;display:inline-block}.hero-row{align-items:flex-start}.hero-logo{width:60px;height:60px}.hero-title{font-size:25px}.hero-sub{font-size:9px;letter-spacing:.14em}.dash-section{padding:16px}.tx-row{gap:8px}.tx-right{font-size:11px}.open-planner{width:100%;text-align:center}.sidebar-brand{padding:14px}}
@media(prefers-reduced-motion:reduce){*,*:before,*:after{animation:none!important;transition:none!important}}
/* UHD 4K rendering helpers */
img{image-rendering:auto;-webkit-font-smoothing:antialiased;text-rendering:geometricPrecision}
html,body,[class*="css"],button,input,textarea,select{ -webkit-font-smoothing:antialiased!important; -moz-osx-font-smoothing:grayscale!important; text-rendering:geometricPrecision!important; }
.stApp,.block-container,section[data-testid="stSidebar"],section[data-testid="stSidebar"] *{ text-rendering:geometricPrecision!important; }

section[data-testid="stSidebar"]>div{padding:18px 12px 28px!important;}
.sidebar-budget-card{margin-top:2px;padding:15px 14px 8px;border:1px solid rgba(114,247,176,.12);border-radius:20px 20px 0 0;background:linear-gradient(145deg,rgba(6,38,23,.72),rgba(2,20,12,.55));box-shadow:inset 0 1px 0 rgba(255,255,255,.08);}
.budget-title{font-family:'Outfit';font-size:12px;font-weight:900;color:#f5fff8;letter-spacing:.02em;}
.sidebar-budget-card + div{margin-top:-1px;}
.sidebar-budget-card + div div[data-baseweb="input"]{border-radius:0 0 15px 15px!important;border-top-color:rgba(114,247,176,.08)!important;}
section[data-testid="stSidebar"] button{font-size:13px!important;letter-spacing:.01em!important;}
section[data-testid="stSidebar"] button p{font-size:13px!important;white-space:nowrap!important;}
@media (min-width:1920px){
  .block-container{max-width:1700px!important;}

  .headbar-title{font-size:31px!important;}
  section[data-testid="stSidebar"] button{min-height:58px!important;}
  section[data-testid="stSidebar"] button p{font-size:14px!important;}
}
.save-img-btn{font-weight:900!important;letter-spacing:.02em!important}

/* FINAL SIDEBAR + DASHBOARD MATCH OVERRIDES */

section[data-testid="stSidebar"]>div{padding:22px 22px 30px!important;}
section[data-testid="stSidebar"] button{
  min-height:54px!important; height:54px!important; margin:7px 0!important;
  padding:0 18px!important; border-radius:18px!important;
  display:flex!important; align-items:center!important; justify-content:flex-start!important;
  text-align:left!important;
  background:linear-gradient(145deg,rgba(9,59,35,.72),rgba(2,30,18,.68))!important;
  border:1px solid rgba(83,236,151,.22)!important;
  box-shadow:0 10px 24px rgba(0,0,0,.20),inset 0 1px 0 rgba(255,255,255,.07)!important;
}
section[data-testid="stSidebar"] button p{
  width:100%!important; font-family:'Manrope',sans-serif!important;
  font-size:12px!important; font-weight:800!important; letter-spacing:.01em!important;
  white-space:nowrap!important; text-align:left!important;
}
section[data-testid="stSidebar"] button:hover{
  transform:translateY(-2px)!important;
  background:linear-gradient(145deg,rgba(13,87,50,.86),rgba(3,38,22,.76))!important;
  border-color:rgba(114,247,176,.55)!important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] h3{
  margin:24px 5px 10px!important; font-size:10px!important;
  letter-spacing:.18em!important;
}
.sidebar-brand{margin-bottom:12px!important;}
.sidebar-live{margin:0 0 10px!important;}
.sidebar-budget-card{border-radius:18px 18px 0 0!important;}
.sidebar-budget-card + div div[data-baseweb="input"]{border-radius:0 0 15px 15px!important;}

/* Remove the old duplicate title-bar language. */
.headbar-container,.headbar-card,.dash-hero{display:none!important;}

/* Single compact centered dashboard heading: no large title bar. */
.dashboard-heading{
  width:100%; display:flex; justify-content:center; align-items:center;
  gap:14px; margin:4px auto 5px; text-align:left;
}
.dashboard-heading img{width:58px;height:58px;object-fit:contain;filter:drop-shadow(0 7px 14px rgba(0,0,0,.28));}
.dashboard-heading-title{font-family:'Outfit',sans-serif;color:#fff;font-size:29px;font-weight:900;letter-spacing:.025em;line-height:1.05;text-align:center;}
.dashboard-heading-sub{margin-top:5px;color:#9fe5b8;font-size:9px;font-weight:800;letter-spacing:.20em;text-align:center;text-transform:uppercase;}
.dashboard-welcome{max-width:920px;margin:0 auto 18px;text-align:center;color:#b9d9c5;font-size:12px;font-weight:600;}
.dashboard-welcome b{color:#72f7b0;}

/* Keep construction ledger entries readable against their light rows. */
.ledger-entry{
    margin:12px 0 8px;
    padding:14px 16px;
    background:#ffffff;
    border:1px solid #d7e3dc;
    border-radius:10px;
    color:#24332a;
}
.ledger-entry p{margin:0;color:#24332a!important;}
.ledger-entry strong{color:#102218!important;}
.ledger-entry hr{display:none;}

/* Re-center the dashboard content after removing both title bars. */
.block-container{max-width:1500px!important;padding-top:18px!important;}
@media (min-width:1920px){

  .block-container{max-width:1500px!important;padding-top:20px!important;}
  .dashboard-heading-title{font-size:32px;}
}
@media(max-width:900px){

  .dashboard-heading-title{font-size:25px;}
  .dashboard-heading img{width:50px;height:50px;}
}
@media(max-width:600px){
  .dashboard-heading{gap:9px;}
  .dashboard-heading-title{font-size:21px;}
  .dashboard-heading-sub{font-size:8px;}
  .dashboard-heading img{width:44px;height:44px;}
}

/* ================================================================
   EDIT HERE #SIDEBAR
   Native Streamlit sidebar controls its own open/closed width.
   Do NOT force a fixed width here.
   ================================================================ */
section[data-testid="stSidebar"] {{
  box-sizing: border-box !important;
  overflow-x: hidden !important;
  min-width: 0 !important;
}}
section[data-testid="stSidebar"] > div,
section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {{
  box-sizing: border-box !important;
  width: 100% !important;
  min-width: 0 !important;
  max-width: 100% !important;
}}
/* Let Streamlit's main area use all space released by the sidebar. */
[data-testid="stAppViewContainer"] > .main,
[data-testid="stAppViewContainer"] .main .block-container {{
  box-sizing: border-box !important;
  max-width: none !important;
}}
/* Never style the collapse button as a normal sidebar menu button. */
section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"] button {{
  transform: none !important;
  min-height: 42px !important;
}}
/* Phone: sidebar is allowed to occupy the screen normally when open. */
@media (max-width: 700px) {{
  section[data-testid="stSidebar"] {{ width: min(88vw, 320px) !important; }}
}}
/* ================================================================
   No overlay logo. Logo/name stay in normal document flow.
   ================================================================ */
}
}
}

</style>
""", unsafe_allow_html=True)

# === FINAL SIDEBAR CLEANUP / NO CLIPPING / NO HORIZONTAL SCROLL ===
st.markdown("""
<style>
/* Keep Streamlit's sidebar stable and readable. */

section[data-testid="stSidebar"] > div {
  width: 100% !important;
  max-width: 100% !important;
  min-width: 0 !important;
  box-sizing: border-box !important;
  padding: 20px 16px 32px !important;
  overflow-x: hidden !important;
  overflow-y: auto !important;
}
section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"],
section[data-testid="stSidebar"] .block-container,
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"],
section[data-testid="stSidebar"] [data-testid="stVerticalBlockBorderWrapper"] {
  width: 100% !important;
  max-width: 100% !important;
  min-width: 0 !important;
  box-sizing: border-box !important;
}
/* Never allow a child to create a wider sidebar. */
section[data-testid="stSidebar"] * {
  box-sizing: border-box !important;
  max-width: 100% !important;
}
section[data-testid="stSidebar"] img,
section[data-testid="stSidebar"] iframe,
section[data-testid="stSidebar"] input,
section[data-testid="stSidebar"] textarea,
section[data-testid="stSidebar"] select {
  max-width: 100% !important;
}
/* Brand: clean, contained and never clipped. */
.sidebar-brand {
  width: 100% !important;
  max-width: 100% !important;
  margin: 0 0 16px !important;
  padding: 18px 16px !important;
  overflow: hidden !important;
}
.brand-row { width:100% !important; min-width:0 !important; }
.brand-logo { flex:0 0 54px !important; width:54px !important; height:54px !important; }
.brand-copy { flex:1 1 auto !important; min-width:0 !important; overflow:hidden !important; }
.brand-title { font-size:16px !important; line-height:1.05 !important; white-space:normal !important; overflow-wrap:anywhere !important; }
.brand-sub { white-space:normal !important; overflow-wrap:anywhere !important; }
.sidebar-live { width:100% !important; white-space:normal !important; overflow-wrap:anywhere !important; text-align:center !important; }
/* Navigation buttons: no forced nowrap, no negative/overflow positioning. */
section[data-testid="stSidebar"] .stButton,
section[data-testid="stSidebar"] .stButton > div { width:100% !important; max-width:100% !important; min-width:0 !important; }
section[data-testid="stSidebar"] .stButton > button:not([data-testid="stSidebarCollapseButton"]):not([aria-label*="Collapse"]):not([aria-label*="Close"]) {
  width:100% !important;
  min-width:0 !important;
  max-width:100% !important;
  min-height:48px !important;
  height:auto !important;
  margin:5px 0 !important;
  padding:12px 14px !important;
  overflow:hidden !important;
  transform:none !important;
}
section[data-testid="stSidebar"] .stButton > button p,
section[data-testid="stSidebar"] .stButton > button span {
  width:auto !important;
  min-width:0 !important;
  max-width:100% !important;
  white-space:normal !important;
  overflow-wrap:anywhere !important;
  text-overflow:clip !important;
  overflow:hidden !important;
  line-height:1.25 !important;
}
section[data-testid="stSidebar"] .stButton > button:hover { transform:translateX(3px) !important; }
/* Budget controls and columns stay inside the sidebar. */
section[data-testid="stSidebar"] [data-testid="stHorizontalBlock"] {
  width:100% !important; max-width:100% !important; min-width:0 !important;
}
section[data-testid="stSidebar"] [data-baseweb="input"],
section[data-testid="stSidebar"] [data-baseweb="select"],
section[data-testid="stSidebar"] .stNumberInput,
section[data-testid="stSidebar"] .stTextInput {
  width:100% !important; max-width:100% !important; min-width:0 !important;
}
/* Native collapse control is independent from navigation styling. */
button[data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"] button {
  width:42px !important; min-width:42px !important; height:42px !important; min-height:42px !important;
  padding:0 !important; margin:10px !important; border-radius:50% !important;
}
/* Desktop: comfortable sidebar. */
@media (min-width: 1400px) {

}
/* Tablet / small laptop. */
@media (min-width: 701px) and (max-width: 1399px) {

}
/* Phone: sidebar may open, but it must remain inside the viewport. */
@media (max-width: 700px) {

  section[data-testid="stSidebar"] > div { padding:16px 12px 28px !important; }
  .sidebar-brand { padding:15px 13px !important; }
  .brand-logo { flex-basis:46px !important; width:46px !important; height:46px !important; }
  .brand-title { font-size:14px !important; }
  section[data-testid="stSidebar"] .stButton > button { min-height:46px !important; padding:10px 12px !important; }
}

/* ================================================================
   EDIT HERE #SIDEBAR
   Native Streamlit sidebar controls its own open/closed width.
   Do NOT force a fixed width here.
   ================================================================ */
section[data-testid="stSidebar"] {{
  box-sizing: border-box !important;
  overflow-x: hidden !important;
  min-width: 0 !important;
}}
section[data-testid="stSidebar"] > div,
section[data-testid="stSidebar"] [data-testid="stSidebarContent"] {{
  box-sizing: border-box !important;
  width: 100% !important;
  min-width: 0 !important;
  max-width: 100% !important;
}}
/* Let Streamlit's main area use all space released by the sidebar. */
[data-testid="stAppViewContainer"] > .main,
[data-testid="stAppViewContainer"] .main .block-container {{
  box-sizing: border-box !important;
  max-width: none !important;
}}
/* Never style the collapse button as a normal sidebar menu button. */
section[data-testid="stSidebar"] [data-testid="stSidebarCollapseButton"],
[data-testid="collapsedControl"] button {{
  transform: none !important;
  min-height: 42px !important;
}}
/* Phone: sidebar is allowed to occupy the screen normally when open. */
@media (max-width: 700px) {{
  section[data-testid="stSidebar"] {{ width: min(88vw, 320px) !important; }}
}}
/* ================================================================
   No overlay logo. Logo/name stay in normal document flow.
   ================================================================ */
}
}
}

/* Keep the same content centered and readable across screen sizes. */
[data-testid="stAppViewContainer"] .main .block-container {
    width: 100% !important;
    max-width: 1500px !important;
    margin-left: auto !important;
    margin-right: auto !important;
    box-sizing: border-box !important;
}
@media (max-width: 700px) {
    [data-testid="stAppViewContainer"] .main .block-container {
        max-width: 100% !important;
        padding: 14px 10px 28px !important;
    }
    [data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important;
    }
    [data-testid="stHorizontalBlock"] > [data-testid="column"] {
        min-width: 100% !important;
        flex: 1 1 100% !important;
    }
}

</style>
""", unsafe_allow_html=True)

if st.session_state.dark_mode:
    st.markdown("""
    <style>
    .stApp, [data-testid="stAppViewContainer"] { background: #101614 !important; color: #edf7ef !important; }
    [data-testid="stHeader"] { background: #101614 !important; }
    [data-testid="stMarkdownContainer"], label, p, h1, h2, h3, h4 { color: #edf7ef !important; }
    [data-baseweb="input"], [data-baseweb="select"] > div, textarea { background: #1b2923 !important; color: #edf7ef !important; }
    </style>
    """, unsafe_allow_html=True)

st.markdown("""
<style>
.photo-scanner-title { color: #f4fff7; font-family: 'Outfit', sans-serif; font-size: 15px; font-weight: 900; letter-spacing: .08em; margin: 14px 0 3px; }
.photo-scanner-subtitle { color: #9fb0a6; font-size: 11px; line-height: 1.4; margin-bottom: 10px; }

/* Final dashboard presentation layer. The project background and sidebar brand remain unchanged. */
.block-container { color: #f7f5ee !important; }
.block-container h1, .block-container h2, .block-container h3 { font-family: 'Outfit', sans-serif !important; letter-spacing: .015em !important; }
.dashboard-heading { margin-bottom: 8px !important; }
.dashboard-heading-title { font-size: 31px !important; letter-spacing: .045em !important; text-shadow: 0 2px 18px rgba(0,0,0,.34); }
.dashboard-heading-sub { color: #8fe0bb !important; letter-spacing: .24em !important; }
.dashboard-welcome { color: #d4e5dc !important; max-width: 760px !important; line-height: 1.6 !important; }
.dashboard-welcome b { color: #ffae8f !important; }
.block-container [data-testid="stMetric"] { min-height: 104px !important; padding: 17px 18px !important; border-radius: 16px !important; background: rgba(12, 38, 29, .82) !important; border: 1px solid rgba(143, 224, 187, .22) !important; box-shadow: 0 12px 28px rgba(0,0,0,.22), inset 0 1px 0 rgba(255,255,255,.09) !important; }
.block-container [data-testid="stMetric"] label { color: #9ed7bd !important; font-size: 9px !important; letter-spacing: .16em !important; }
.block-container [data-testid="stMetricValue"] { color: #fffaf2 !important; font-family: 'Outfit', sans-serif !important; font-size: 26px !important; }
.dash-section { border-radius: 16px !important; background: rgba(10, 30, 24, .78) !important; border-color: rgba(143, 224, 187, .18) !important; box-shadow: 0 16px 34px rgba(0,0,0,.24) !important; }
.section-title { color: #fffaf2 !important; font-size: 15px !important; letter-spacing: .075em !important; }
.section-head span { color: #9ed7bd !important; }
.tx-row { border-bottom-color: rgba(211, 238, 222, .10) !important; }
.tx-name { color: #fffaf2 !important; font-family: 'Outfit', sans-serif !important; }
.tx-type, .tx-date, .schedule-muted { color: #9db9aa !important; }
.tx-right { color: #ffae8f !important; }
.schedule-title { color: #fffaf2 !important; font-size: 15px !important; letter-spacing: .055em !important; }
@media (max-width: 600px) {
    .dashboard-heading-title { font-size: 22px !important; }
    .dashboard-heading-sub { letter-spacing: .14em !important; }
    .block-container [data-testid="stMetric"] { min-height: 88px !important; }
}
</style>
""", unsafe_allow_html=True)

with st.sidebar:
    st.markdown(f"""
    <div class="sidebar-brand">
      <div class="brand-row">
        <img class="brand-logo" src="{AILYN_LOGO_DATA}" alt="Ailyn Construction Logo">
        <div class="brand-copy">
        <div class="brand-title"><span>AILYN HOUSE</span><span>PROJECT</span></div>
          <div class="brand-sub">Official Project Control</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown(
        f"<div class='sidebar-live'><span>●</span> &nbsp; LIVE SYSTEM &nbsp; • &nbsp; "
        f"{manila_now().strftime('%I:%M %p  |  %b %d')}</div>",
        unsafe_allow_html=True
    )

    st.markdown(
        "<div class='photo-scanner-title'>PHOTO SCANNER</div>"
        "<div class='photo-scanner-subtitle'>Capture receipts and project progress</div>",
        unsafe_allow_html=True,
    )
    if st.button("📷 TAKE PHOTO", use_container_width=True, key="take_photo_sidebar"):
        set_view("photo_scanner")

    if st.button("📝 NOTES", use_container_width=True, key="sidebar_notes_popup"):
        notes_dialog()

    st.subheader("Executive Overview")
    if st.button("📊   Dashboard   ›", use_container_width=True, key="side_dashboard"):
        set_view("home")

    st.markdown("<div class='sidebar-budget-card'><div class='budget-title'>Budget Control</div></div>",
                unsafe_allow_html=True)
    if st.button("💰   Apply Budget", use_container_width=True, key="side_budget"):
        budget_dialog()

    if st.button("🔄   Restart System   ›", use_container_width=True, key="side_restart"):
        clear_all()
        set_view("home")

    with st.expander("Project Details", expanded=False):
        project = st.session_state.project
        st.caption(f"{project.get('status', 'Active')} project")
        if st.button("EDIT PROJECT DETAILS", use_container_width=True, key="side_project_details"):
            project_settings_dialog()

    st.subheader("Project Control")
    if st.button("📝   New Work Entry   ›", use_container_width=True, key="side_new_work"):
        set_view("planner_input")
    if st.button("📅   Schedule & Progress   ›", use_container_width=True, key="side_schedule"):
        set_view("planner_output")
    if st.button("🧰   Project Tools   ›", use_container_width=True, key="side_project_tools"):
        set_view("project_tools")
    if st.button("⚙️   Settings   ›", use_container_width=True, key="side_settings"):
        settings_dialog()

    st.subheader("Financial Operations")
    if st.button("🧱   Material Entry   ›", use_container_width=True, key="side_material"):
        set_view("material")
    if st.button("🧾   Expense Entry   ›", use_container_width=True, key="side_expense"):
        set_view("expense")
    if st.button("🏦   Encash Deposit   ›", use_container_width=True, key="side_excess"):
        set_view("excess")
    if st.button("📒   Financial Ledger   ›", use_container_width=True, key="side_ledger"):
        set_view("ledger")
    if st.button("📈   Financial Report   ›", use_container_width=True, key="side_financial_report"):
        set_view("export")

    st.subheader("Payroll Operations")
    if st.button("📊   Payroll Dashboard   ›", use_container_width=True, key="side_payroll_dashboard"):
        set_view("payroll_dashboard")
    if st.button("👷   Labor Account   ›", use_container_width=True, key="side_labor"):
        set_view("add_labor")
    if st.button("💳   Payroll Expense   ›", use_container_width=True, key="side_payroll_expense"):
        set_view("add_payroll_expense")
    if st.button("🪙   Account Remainder   ›", use_container_width=True, key="side_payroll_remaining"):
        set_view("payroll_remaining")
    if st.button("👥   Labor Accounts   ›", use_container_width=True, key="side_payroll_ledger"):
        set_view("payroll_ledger")
    if st.button("📋   Payroll Report   ›", use_container_width=True, key="side_payroll_report"):
        set_view("payroll_export")
    if st.button("🗃️   Receipts Archive   ›", use_container_width=True, key="side_archive"):
        set_view("receipt_archive")

    st.subheader("Administrator")
    if st.button("🔐   Admin Console   ›", use_container_width=True, key="side_admin"):
        set_view("update")

view = st.session_state.view

if view == "home":
    budget = float(st.session_state.budget or 0)
    used = float(get_total() or 0)
    balance = float(get_balance() or 0)
    material = float(total_materials() or 0)
    expenses = float(total_expenses() or 0)
    excess = float(total_excess() or 0)
    chart_total = max(material + expenses + excess, 1.0)
    p1 = material / chart_total * 360
    p2 = p1 + expenses / chart_total * 360
    p3 = p2 + excess / chart_total * 360
    today_key = manila_now().strftime("%Y-%m-%d")
    today_tasks = [t for t in st.session_state.planner_tasks if t.get("date_obj") == today_key]
    upcoming_tasks = [t for t in st.session_state.planner_tasks if t.get("date_obj", "") >= today_key]

    st.markdown(f"""
    <div class="dashboard-heading">
    <img src="{AILYN_LOGO_DATA}" alt="Ailyn Construction Logo">
      <div>
        <div class="dashboard-heading-title">AILYN HOUSE PROJECT</div>
        <div class="dashboard-heading-sub">PROJECT MANAGEMENT SYSTEM</div>
      </div>
    </div>
    <div class="dashboard-welcome">🛡️ &nbsp; Welcome back, <b>{st.session_state.project.get("name", "Ailyn House Project")}</b> &nbsp;|&nbsp; Manage your construction project efficiently.</div>
    """, unsafe_allow_html=True)
    project = st.session_state.project
    if project.get("client") or project.get("address") or project.get("target_date"):
        st.caption(
            f"Client: {project.get('client') or 'Not set'}  |  Site: {project.get('address') or 'Not set'}  |  "
            f"Target: {project.get('target_date') or 'Not set'}  |  Status: {project.get('status', 'Active')}"
        )
    overdue_tasks = [
        task for task in st.session_state.planner_tasks
        if task.get("date_obj", "") < manila_now().strftime("%Y-%m-%d")
        and task.get("status") != "Completed"
    ]
    if balance < 0:
        st.error(f"Budget warning: project is over budget by PHP {abs(balance):,.2f}.")
    if overdue_tasks:
        st.warning(f"{len(overdue_tasks)} scheduled task(s) are overdue.")

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("TOTAL BUDGET", f"₱{budget:,.2f}")
    with m2:
        st.metric("TOTAL EXPENSES", f"₱{used:,.2f}")
    with m3:
        st.metric("REMAINING BALANCE", f"₱{balance:,.2f}")
    with m4:
        st.metric(
            f"PROJECT SPENT THIS MONTH ({manila_now().strftime('%b %Y').upper()})",
            f"₱{monthly_construction_spend():,.2f}",
        )

    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
    left, right = st.columns([1.05, 1])
    with left:
        st.markdown(f"""
        <div class="dash-section">
          <div class="section-head"><div class="section-title" style="margin:0">EXPENSES OVERVIEW</div><span style="font-size:11px;color:#7b867f;font-weight:700">THIS PROJECT</span></div>
          <div class="donut-wrap">
            <div class="donut" style="--p1:{p1}deg;--p2:{p2}deg;--p3:{p3}deg"><div class="donut-center">₱{used:,.0f}<small>Total Expenses</small></div></div>
            <div class="legend">
              <div class="legend-row"><span><i class="dot" style="background:#075c28"></i>Materials</span><b>₱{material:,.2f}</b></div>
              <div class="legend-row"><span><i class="dot" style="background:#e0aa25"></i>Expenses</span><b>₱{expenses:,.2f}</b></div>
              <div class="legend-row"><span><i class="dot" style="background:#e85d4a"></i>Excess</span><b>₱{excess:,.2f}</b></div>
            </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
    with right:
        tx = list(reversed(st.session_state.records))[:5]
        tx_html = ""
        if tx:
            for r in tx:
                icon = "🛒" if r.get("type") == "material" else "▣" if r.get("type") == "expense" else "+"
                tx_html += f'''<div class="tx-row"><div class="tx-left"><div class="tx-icon">{icon}</div><div><div class="tx-name">{r.get("name", "Transaction")}</div><div class="tx-type">{str(r.get("type", "")).title()}</div></div></div><div class="tx-right">₱{float(r.get("amount", 0)):,.2f}<div class="tx-date">{r.get("date", "")}</div></div></div>'''
        else:
            tx_html = '<div style="padding:30px 0;color:#7a857e;text-align:center">No transactions yet.</div>'
        st.markdown(
            f'''<div class="dash-section"><div class="section-head"><div class="section-title" style="margin:0">RECENT TRANSACTIONS</div><span style="font-size:11px;color:#7b867f">LATEST 5</span></div>{tx_html}</div>''',
            unsafe_allow_html=True)

    st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class="dash-section">
      <div class="schedule">
        <div class="schedule-icon">▦</div>
        <div><div class="schedule-title">TODAY'S SCHEDULE</div><div style="font-weight:800;font-size:13px;margin-top:4px">{manila_now().strftime('%B %d, %Y (%A)')}</div><div class="schedule-muted">{len(today_tasks)} task(s) scheduled for today.</div></div>
        <div style="width:1px;height:58px;background:#dfe8e1;margin:0 12px"></div>
        <div><div class="schedule-title">UPCOMING TASKS</div><div style="font-weight:800;font-size:13px;margin-top:4px">{len(upcoming_tasks)} task(s) planned</div><div class="schedule-muted">Stay on track and manage your construction tasks.</div></div>
        <div style="margin-left:auto"><div class="open-planner">▣ &nbsp; Open Planner</div></div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    if st.button("OPEN CONSTRUCTION PLANNER", use_container_width=True):
        set_view("planner_output")

elif view == "payroll_dashboard":
    payroll_labor = sum(float(record.get("net", 0)) for record in st.session_state.labor_records)
    payroll_expenses = sum(float(record.get("price", 0)) for record in st.session_state.payroll_expenses)
    payroll_total = payroll_labor + payroll_expenses
    payroll_remaining = float(st.session_state.remaining_money or 0)
    payroll_output = payroll_total - payroll_remaining

    st.markdown(f"""
        <div class="dashboard-heading">
            <img src="{AILYN_LOGO_DATA}" alt="Ailyn Construction Logo">
            <div>
                <div class="dashboard-heading-title">PAYROLL DASHBOARD</div>
                <div class="dashboard-heading-sub">AILYN HOUSE PROJECT | AILYN HOUSE</div>
            </div>
        </div>
        <div class="dashboard-welcome">Payroll control center for labor accounts, payroll expenses, and final payouts.</div>
    """, unsafe_allow_html=True)
    st.caption("A focused view of labor, payroll expenses, and the current payout.")
    p1, p2, p3, p4 = st.columns(4)
    with p1:
        st.metric("Workers", len(st.session_state.labor_records))
    with p2:
        st.metric("Net labor", f"PHP {payroll_labor:,.2f}")
    with p3:
        st.metric("Payroll expenses", f"PHP {payroll_expenses:,.2f}")
    with p4:
        st.metric("Final payout", f"PHP {payroll_output:,.2f}")

    st.divider()
    left, right = st.columns(2)
    with left:
        st.subheader("Role summary")
        role_rows = []
        for role in FULL_DAY_RATES:
            role_records = [r for r in st.session_state.labor_records if r.get("role") == role]
            role_rows.append({
                "Role": role,
                "Workers": len(role_records),
                "Days": sum(float(r.get("days", 0)) for r in role_records),
                "Net pay": f"PHP {sum(float(r.get('net', 0)) for r in role_records):,.2f}",
            })
        st.dataframe(role_rows, use_container_width=True, hide_index=True)
    with right:
        st.subheader("Payroll controls")
        st.write(f"Current remainder: **PHP {payroll_remaining:,.2f}**")
        if st.button("ADD LABOR ACCOUNT", use_container_width=True):
            payroll_labor_dialog()
        if st.button("OPEN PAYROLL LEDGER", use_container_width=True):
            payroll_ledger_dialog()
        if st.button("CREATE PAYROLL REPORT", use_container_width=True):
            payroll_report_dialog()

    st.subheader("Latest labor accounts")
    if st.session_state.labor_records:
        st.dataframe([
            {"Worker": r.get("name", ""), "Role": r.get("role", "Labor"),
             "Days": f"{float(r.get('days', 0)):.1f}",
             "Net pay": f"PHP {float(r.get('net', 0)):,.2f}"}
            for r in reversed(st.session_state.labor_records[-8:])
        ], use_container_width=True, hide_index=True)
    else:
        st.info("No labor accounts yet. Add the first worker from Payroll Operations.")

elif view == "planner_input":
    st.subheader("📅 PLANNER INPUT - ADD NEW WORK TASK")
    st.caption("Select date details, work description, and optional photo proofs.")
    with st.form(key="planner_input_form", clear_on_submit=True):
        selected_date = st.date_input("Select Day, Month, and Year", value=manila_now().date())
        work_description = st.text_area("Work Description / Task Details", placeholder="Describe construction work...")
        phase = st.selectbox("Construction Phase",
                             ["Site Prep", "Foundation", "Framing & Masonry", "Roofing", "Plumbing & Electrical",
                              "Finishing", "Inspection"])
        uploaded_files = st.file_uploader("Upload Work Proof Photos (Optional)", type=["jpg", "jpeg", "png"],
                                          accept_multiple_files=True)
        submitted = st.form_submit_button("💾 SAVE TASK TO PERMANENT STORAGE")

        if submitted:
            if work_description.strip():
                photos_base64 = []
                if uploaded_files:
                    for file in uploaded_files:
                        bytes_data = file.read()
                        b64_str = base64.b64encode(bytes_data).decode('utf-8')
                        mime_type = file.type or "image/png"
                        photos_base64.append(f"data:{mime_type};base64,{b64_str}")
                st.session_state.planner_tasks.append({
                    "id": str(time.time()),
                    "day": selected_date.strftime("%d"),
                    "month": selected_date.strftime("%B"),
                    "year": selected_date.strftime("%Y"),
                    "date_obj": selected_date.strftime("%Y-%m-%d"),
                    "name": work_description.upper(),
                    "phase": phase,
                    "status": "Not Started",
                    "photos": photos_base64
                })
                persist_state()
                st.success("Task & photos permanently saved!")
                st.rerun()
            else:
                st.warning("Please fill in the work description.")
    st.divider()
    if st.button("🏠 RETURN TO HOME", use_container_width=True):
        set_view("home")

elif view == "planner_output":
    st.subheader("📋 PLANNER OUTPUT - WORK SCHEDULE CALENDAR")
    tasks = st.session_state.planner_tasks
    if not tasks:
        st.info("No work scheduled yet.")
    else:
        sorted_tasks = sorted(tasks, key=lambda x: x.get('date_obj', ''))
        cards_html = '<div class="cal-grid">'
        for t in sorted_tasks:
            badge_class = "badge-completed" if t['status'] == "Completed" else "badge-inprogress" if t[
                                                                                                         'status'] == "In Progress" else "badge-notstarted"
            photos_thumbs = ""
            if t.get("photos"):
                photos_thumbs = '<div class="card-photos">'
                for p in t["photos"]:
                    photos_thumbs += f'<img src="{p}" class="card-photo-thumb" />'
                photos_thumbs += '</div>'
            cards_html += f"""
            <div class="cal-card">
              <div class="cal-date-badge">📅 {t.get("month", "")} {t.get("day", "")}, {t.get("year", "")}</div>
              <div class="cal-task-title">{t["name"]}</div>
              <div class="cal-phase">🛠️ Phase: {t["phase"]}</div>
              <div class="cal-status-tag {badge_class}">{t["status"]}</div>
              {photos_thumbs}
            </div>
            """
        cards_html += '</div>'
        st.markdown(cards_html, unsafe_allow_html=True)
        st.markdown("<div class='sidebar-gap'></div>", unsafe_allow_html=True)
        st.subheader("⚙️ Task Management & Photo Inspector")

        for t in list(sorted_tasks):
            with st.expander(f"📌 {t.get('month')} {t.get('day')}, {t.get('year')} - {t['name']} ({t['status']})",
                             expanded=False):
                col1, col2 = st.columns([2, 1])
                with col1:
                    new_status = st.selectbox("Update Status", ["Not Started", "In Progress", "Completed"],
                                              index=["Not Started", "In Progress", "Completed"].index(t["status"]),
                                              key=f"st_{t['id']}")
                    if new_status != t["status"]:
                        t["status"] = new_status
                        persist_state()
                        st.rerun()
                with col2:
                    if st.button("❌ Delete Task", key=f"del_{t['id']}", use_container_width=True):
                        st.session_state.planner_tasks = [x for x in st.session_state.planner_tasks if
                                                          x["id"] != t["id"]]
                        persist_state()
                        st.rerun()

                st.markdown("#### 📸 Work Gallery for this Day")
                if t.get("photos"):
                    img_cols = st.columns(4)
                    for idx, photo_b64 in enumerate(list(t["photos"])):
                        with img_cols[idx % 4]:
                            st.image(photo_b64, use_container_width=True)
                            if st.button("🗑️ Remove Photo", key=f"del_img_{t['id']}_{idx}", use_container_width=True):
                                t["photos"].pop(idx)
                                persist_state()
                                st.rerun()
                else:
                    st.info("No photo proof attached for this work day yet.")

                st.markdown("##### Add More Photos")
                with st.form(key=f"upload_form_{t['id']}", clear_on_submit=True):
                    new_photos = st.file_uploader("Upload Additional Photos", type=["jpg", "jpeg", "png"],
                                                  accept_multiple_files=True, key=f"up_{t['id']}")
                    add_photos_btn = st.form_submit_button("📤 UPLOAD PHOTOS")
                    if add_photos_btn and new_photos:
                        if "photos" not in t or t["photos"] is None:
                            t["photos"] = []
                        for f in new_photos:
                            bytes_data = f.read()
                            b64_str = base64.b64encode(bytes_data).decode('utf-8')
                            mime_type = f.type or "image/png"
                            t["photos"].append(f"data:{mime_type};base64,{b64_str}")
                        persist_state()
                        st.success("Photos added successfully!")
                        st.rerun()

    st.divider()
    if st.button("🏠 RETURN TO HOME", use_container_width=True):
        set_view("home")

elif view == "material":
    st.subheader("➕ ADD MATERIAL")
    with st.form(key="material_form", clear_on_submit=True):
        name = st.text_input("Material Name", key="material_name")
        price = st.number_input("Price", min_value=0.01, value=None, placeholder="0.00", key="material_price")
        qty = st.number_input("Qty", min_value=1, value=None, placeholder="1", key="material_qty")
        delivery = st.number_input("Delivery", min_value=0.0, value=None, placeholder="0.00", key="material_delivery")
        sender = st.selectbox("Sender", ["Garr", "Aily"])
        submitted = st.form_submit_button(label="SAVE MATERIAL")
        if submitted:
            ok = add_tx(name, price, qty, delivery or 0.0, "material", sender)
            if ok:
                st.success("Saved! Ready for next order.")
                st.rerun()
            else:
                st.warning("Invalid data, please fill out Price and Qty.")
    st.divider()
    if st.button("🏠 RETURN TO HOME", use_container_width=True):
        set_view("home")

elif view == "expense":
    st.subheader("➕ ADD CONSTRUCTION EXPENSE")
    with st.form(key="expense_form", clear_on_submit=True):
        name = st.text_input("Expense Name")
        amount = st.number_input("Amount", min_value=0.01, value=None, placeholder="0.00")
        sender = st.selectbox("Sender", ["Garr", "Aily"])
        submitted = st.form_submit_button(label="SAVE EXPENSE")
        if submitted:
            if amount and amount > 0:
                add_tx(name, amount, 1, 0, "expense", sender)
                st.success("Expense Added → Ledger Updated")
                st.rerun()
            else:
                st.warning("Please enter an amount greater than zero.")
    st.divider()
    if st.button("🏠 RETURN TO HOME", use_container_width=True):
        set_view("home")

elif view == "excess":
    st.subheader("➕ ADD EXCESS MONEY")
    with st.form(key="excess_form", clear_on_submit=True):
        name = st.text_input("Reason")
        amount = st.number_input("Amount", min_value=0.01, value=None, placeholder="0.00")
        sender = st.selectbox("Sender", ["Garr", "Aily"])
        submitted = st.form_submit_button(label="ADD EXCESS")
        if submitted:
            if amount and amount > 0:
                st.session_state.records.append({
                    "id": str(time.time()),
                    "date": manila_now().strftime("%b %d, %Y"),
                    "recorded_at": manila_now().isoformat(),
                    "name": name.upper(),
                    "price": float(amount),
                    "qty": 1,
                    "delivery": 0.0,
                    "amount": float(amount),
                    "type": "excess",
                    "sender": sender
                })
                persist_state()
                st.success("Excess Added")
                st.rerun()
            else:
                st.warning("Please enter a valid amount.")
    st.divider()
    if st.button("🏠 RETURN TO HOME", use_container_width=True):
        set_view("home")

elif view == "ledger":
    st.subheader("📖 CONSTRUCTION LEDGER")
    ledger_query = st.text_input("Search construction entries", key="construction_ledger_search").strip().lower()
    visible_records = [
        record for record in st.session_state.records
        if not ledger_query or ledger_query in str(record).lower()
    ]
    if not st.session_state.records:
        st.info("No transaction records found in ledger.")
    elif not visible_records:
        st.info("No construction entries match your search.")
    else:
        for r in visible_records:
            if st.session_state.editing_record_id == r["id"]:
                with st.form(key=f"edit_record_form_{r['id']}"):
                    edited_name = st.text_input("Name", value=r.get("name", ""))
                    edited_amount = st.number_input(
                        "Amount / Unit Price",
                        min_value=0.01,
                        value=float(r.get("price", r.get("amount", 0.01))),
                    )
                    edited_qty = st.number_input(
                        "Quantity",
                        min_value=1,
                        value=int(r.get("qty", 1)),
                        step=1,
                        disabled=r.get("type") != "material",
                    )
                    edited_delivery = st.number_input(
                        "Delivery",
                        min_value=0.0,
                        value=float(r.get("delivery", 0.0)),
                        disabled=r.get("type") != "material",
                    )
                    edited_sender = st.selectbox(
                        "Sender",
                        ["Garr", "Aily"],
                        index=0 if r.get("sender") == "Garr" else 1,
                    )
                    save_record = st.form_submit_button("SAVE EDIT")
                    cancel_record = st.form_submit_button("CANCEL")
                if save_record:
                    if edited_name.strip() and edited_amount > 0:
                        r["name"] = edited_name.strip().upper()
                        r["price"] = float(edited_amount)
                        r["qty"] = int(edited_qty)
                        r["delivery"] = float(edited_delivery)
                        r["sender"] = edited_sender
                        r["amount"] = (
                            float(edited_amount) * int(edited_qty) + float(edited_delivery)
                            if r.get("type") == "material"
                            else float(edited_amount)
                        )
                        st.session_state.editing_record_id = None
                        persist_state()
                        st.rerun()
                    else:
                        st.warning("Please enter a name and a valid amount.")
                if cancel_record:
                    st.session_state.editing_record_id = None
                    st.rerun()
                continue

            st.markdown(f"""
            <div class="ledger-entry">
            <strong>{r['name']}</strong> • PHP {float(r['amount']):,.2f}<br>
            <span>👤 {r['sender']} | 🏷️ {r['type']} | 📅 {r['date']}</span>
            </div>
            """, unsafe_allow_html=True)
            if st.button("✏️ EDIT ENTRY", key=f"edit_{r['id']}", use_container_width=True):
                st.session_state.editing_record_id = r["id"]
                st.rerun()
            if st.button("❌ DELETE ENTRY", key=f"del_{r['id']}", use_container_width=True):
                st.session_state.records = [x for x in st.session_state.records if x["id"] != r["id"]]
                persist_state()
                st.rerun()

elif view == "add_labor":
    st.subheader("👷 ADD LABOR ACCOUNT")
    st.caption("Click a role button below (Cashier POS Style) to select the work role quickly:")

    col_r1, col_r2, col_r3 = st.columns(3)
    with col_r1:
        if st.button("LABOR\n₱500 / day", use_container_width=True):
            st.session_state.selected_role = "Labor"
    with col_r2:
        if st.button("SKILL\n₱650 / day", use_container_width=True):
            st.session_state.selected_role = "Skill"
    with col_r3:
        if st.button("FORMAN\n₱800 / day", use_container_width=True):
            st.session_state.selected_role = "Forman"

    active_role = st.session_state.selected_role
    st.markdown(f"""
    <div class="pos-role-box">
      CURRENTLY SELECTED ROLE: <span style="color:#4ade80; font-size:20px;">{active_role.upper()}</span> (₱{FULL_DAY_RATES[active_role]:,.2f}/day)
    </div>
    """, unsafe_allow_html=True)

    with st.form(key="labor_input_form", clear_on_submit=True):
        name = st.text_input("Worker Name")
        days = st.number_input("Worked Days / Point", min_value=0.0, value=None, placeholder="Enter days")
        ca = st.number_input("Cash Advance (C.A.)", min_value=0.0, value=None, placeholder="0.00")
        submitted = st.form_submit_button("💾 SAVE LABOR ACCOUNT")

        if submitted:
            d = float(days or 0.0)
            c = float(ca or 0.0)
            if d > 0 and name.strip():
                gross_pay, full_pay, partial_pay = calculate_labor_pay(d, active_role)
                net = gross_pay - c
                rate = FULL_DAY_RATES.get(active_role, 0.0)
                st.session_state.labor_records.append({
                    "id": str(uuid.uuid4()),
                    "date": manila_now().strftime("%b %d, %Y"),
                    "recorded_at": manila_now().isoformat(),
                    "name": name.upper(),
                    "role": active_role,
                    "days": d,
                    "rate": rate,
                    "gross_pay": gross_pay,
                    "ca": c,
                    "net": net
                })
                persist_state()
                st.success(f"Record for {name.upper()} ({active_role}, {d:.1f} day) added. Net: PHP {net:,.2f}")
                st.rerun()
            else:
                st.warning("Please enter a worker name and valid worked days/points.")

elif view == "add_payroll_expense":
    st.subheader("➕ ADD PAYROLL EXPENSE")
    with st.form(key="payroll_expense_form", clear_on_submit=True):
        desc = st.text_input("Expense Description")
        amt = st.number_input("Amount", min_value=0.01, value=None, placeholder="0.00")
        submitted = st.form_submit_button("💾 SAVE EXPENSE")
        if submitted:
            if amt and amt > 0:
                st.session_state.payroll_expenses.append({
                    "id": str(uuid.uuid4()),
                    "date": manila_now().strftime("%b %d, %Y"),
                    "recorded_at": manila_now().isoformat(),
                    "item": desc.upper(),
                    "price": float(amt)
                })
                persist_state()
                st.success(f"Expense {desc.upper()} added.")
                st.rerun()
            else:
                st.warning("Please enter a valid amount.")

elif view == "payroll_remaining":
    st.subheader("⚙️ SET REMAINING MONEY")
    res = st.number_input("Leftover/Remaining money to subtract from total", min_value=0.0, value=None,
                          placeholder="0.00")
    if st.button("APPLY REMAINING MONEY", use_container_width=True):
        if res is not None:
            st.session_state.remaining_money = float(res)
            persist_state()
            st.success("Remaining money applied.")
            st.rerun()
        else:
            st.warning("Please enter an amount.")

elif view == "payroll_ledger":
    st.subheader("📋 LABOR & PAYROLL LEDGER")
    payroll_query = st.text_input("Search payroll entries", key="payroll_ledger_search").strip().lower()
    visible_labor = [record for record in st.session_state.labor_records if not payroll_query or payroll_query in str(record).lower()]
    visible_payroll_expenses = [record for record in st.session_state.payroll_expenses if not payroll_query or payroll_query in str(record).lower()]
    st.markdown("### Labor Records")
    if not st.session_state.labor_records:
        st.info("No labor records.")
    else:
        for i, r in enumerate(st.session_state.labor_records):
            if r not in visible_labor:
                continue
            if st.session_state.editing_labor_index == i:
                with st.form(key=f"edit_labor_form_{i}"):
                    edited_worker = st.text_input("Worker Name", value=r.get("name", ""))
                    edited_role = st.selectbox(
                        "Role",
                        list(FULL_DAY_RATES),
                        index=list(FULL_DAY_RATES).index(r.get("role", "Labor")),
                    )
                    edited_days = st.number_input(
                        "Worked Days / Point",
                        min_value=0.0,
                        value=float(r.get("days", 1.0)),
                    )
                    edited_ca = st.number_input(
                        "Cash Advance (C.A.)",
                        min_value=0.0,
                        value=float(r.get("ca", 0.0)),
                    )
                    save_labor = st.form_submit_button("SAVE EDIT")
                    cancel_labor = st.form_submit_button("CANCEL")
                if save_labor:
                    if edited_worker.strip() and edited_days > 0:
                        gross_pay, full_pay, partial_pay = calculate_labor_pay(float(edited_days), edited_role)
                        r.update({
                            "name": edited_worker.strip().upper(),
                            "role": edited_role,
                            "days": float(edited_days),
                            "rate": FULL_DAY_RATES[edited_role],
                            "gross_pay": gross_pay,
                            "full_pay": full_pay,
                            "partial_pay": partial_pay,
                            "ca": float(edited_ca),
                            "net": gross_pay - float(edited_ca),
                        })
                        st.session_state.editing_labor_index = None
                        persist_state()
                        st.rerun()
                    else:
                        st.warning("Please enter a worker name and valid worked days.")
                if cancel_labor:
                    st.session_state.editing_labor_index = None
                    st.rerun()
                continue

            role_disp = r.get('role', 'Labor')
            gross_disp = r.get('gross_pay', r['days'] * r['rate'])
            st.markdown(f"""
            ---
            **{r['name']}** ({role_disp}) • Worked: {r['days']:.1f} Day(s)  
            • Gross Pay: PHP {gross_disp:,.2f}  
            • C.A.: PHP {r['ca']:,.2f}  
            • **Net Pay: PHP {r['net']:,.2f}**
            """)
            if st.button("✏️ EDIT LABOR ENTRY", key=f"edit_lab_{i}", use_container_width=True):
                st.session_state.editing_labor_index = i
                st.rerun()
            if st.button("❌ DELETE LABOR ENTRY", key=f"del_lab_{i}", use_container_width=True):
                st.session_state.labor_records.pop(i)
                persist_state()
                st.rerun()

    st.markdown("<div class='sidebar-gap'></div>", unsafe_allow_html=True)
    st.markdown("### Payroll Expenses")
    if not st.session_state.payroll_expenses:
        st.info("No payroll expenses.")
    else:
        for i, e in enumerate(st.session_state.payroll_expenses):
            if e not in visible_payroll_expenses:
                continue
            if st.session_state.editing_payroll_expense_index == i:
                with st.form(key=f"edit_payroll_expense_form_{i}"):
                    edited_description = st.text_input("Expense Description", value=e.get("item", ""))
                    edited_price = st.number_input(
                        "Amount",
                        min_value=0.01,
                        value=float(e.get("price", 0.01)),
                    )
                    save_payroll_expense = st.form_submit_button("SAVE EDIT")
                    cancel_payroll_expense = st.form_submit_button("CANCEL")
                if save_payroll_expense:
                    if edited_description.strip() and edited_price > 0:
                        e["item"] = edited_description.strip().upper()
                        e["price"] = float(edited_price)
                        st.session_state.editing_payroll_expense_index = None
                        persist_state()
                        st.rerun()
                    else:
                        st.warning("Please enter a description and valid amount.")
                if cancel_payroll_expense:
                    st.session_state.editing_payroll_expense_index = None
                    st.rerun()
                continue

            st.markdown(f"- **{e['item']}**: PHP {e['price']:,.2f}")
            if st.button("✏️ EDIT PAYROLL EXPENSE", key=f"edit_pay_exp_{i}", use_container_width=True):
                st.session_state.editing_payroll_expense_index = i
                st.rerun()
            if st.button("❌ DELETE PAYROLL EXPENSE", key=f"del_pay_exp_{i}", use_container_width=True):
                st.session_state.payroll_expenses.pop(i)
                persist_state()
                st.rerun()

elif view == "export":
    st.subheader("📄 EXPORT CONSTRUCTION REPORT")
    receipt_title = st.text_input("Receipt Title", value="OFFICIAL RECEIPT",
                                  placeholder="Enter a title for this receipt")
    html = build_html_report(st.session_state.records, st.session_state.budget, custom_title=receipt_title)
    st.components.v1.html(
        html,
        height=receipt_preview_height(len(st.session_state.records), row_height=35, base_height=540),
        scrolling=True,
    )
    if st.button("💾 SAVE RECEIPT TO ARCHIVE", use_container_width=True):
        if receipt_title.strip():
            archive_path = save_report_html("construction", html, title=receipt_title)
            st.success(f"Saved to archive: {archive_path}")
        else:
            st.warning("Please enter a title before saving.")
    st.download_button(
        label="📥 DOWNLOAD CONSTRUCTION REPORT HTML",
        data=html,
        file_name="construction_report.html",
        mime="text/html",
        use_container_width=True
    )
    if st.button("📂 OPEN RECEIPT ARCHIVE", use_container_width=True):
        set_view("receipt_archive")

elif view == "payroll_export":
    st.subheader("📄 EXPORT PAYROLL REPORT")
    receipt_title = st.text_input("Receipt Title", value="OFFICIAL LABOR TALLY",
                                  placeholder="Enter a title for this receipt")
    html, total = generate_payroll_html(
        st.session_state.labor_records,
        st.session_state.payroll_expenses,
        st.session_state.remaining_money,
        custom_title=receipt_title
    )
    st.components.v1.html(
        html,
        height=receipt_preview_height(
            len(st.session_state.labor_records) + len(st.session_state.payroll_expenses),
            row_height=58,
        ),
        scrolling=True,
    )
    if st.button("💾 SAVE RECEIPT TO ARCHIVE", use_container_width=True):
        if receipt_title.strip():
            archive_path = save_report_html("payroll", html, title=receipt_title)
            st.success(f"Saved to archive: {archive_path}")
        else:
            st.warning("Please enter a title before saving.")
    st.download_button(
        label="📥 DOWNLOAD PAYROLL REPORT HTML",
        data=html,
        file_name="payroll_report.html",
        mime="text/html",
        use_container_width=True
    )
    if st.button("📂 OPEN RECEIPT ARCHIVE", use_container_width=True):
        set_view("receipt_archive")
    if st.button("📧 EMAIL PAYROLL REPORT", use_container_width=True):
        try:
            if not SENDER_EMAIL or not SENDER_PASSWORD:
                raise ValueError("Email is disabled. Configure AILYN_SENDER_EMAIL and AILYN_SENDER_PASSWORD.")
            msg = EmailMessage()
            msg['Subject'] = f"Construction Report: PHP {total:,.2f} - {manila_now().strftime('%Y-%m-%d')}"
            msg['From'] = SENDER_EMAIL
            msg['To'] = RECEIVER_EMAIL
            msg.add_alternative(html, subtype='html')
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.login(SENDER_EMAIL, SENDER_PASSWORD)
                smtp.send_message(msg)
            st.success("✅ SUCCESS! Emailed report.")
        except Exception as e:
            st.error(f"❌ EMAIL FAILED: {e}")

elif view == "receipt_archive":
    st.subheader("📂 RECEIPT ARCHIVE | AILYN HOUSE")
    st.caption("Every saved receipt is preserved here and summarized in the complete Excel ledger.")
    st.metric(
        f"TOTAL SPENT THIS MONTH ({manila_now().strftime('%b %Y').upper()})",
        f"PHP {monthly_spend():,.2f}",
    )
    backup_col, restore_col = st.columns(2)
    with backup_col:
        if st.button("CREATE DATABASE BACKUP", use_container_width=True):
            backup_path = create_backup()
            with open(backup_path, "rb") as backup_file:
                st.download_button(
                    "DOWNLOAD BACKUP",
                    data=backup_file.read(),
                    file_name=os.path.basename(backup_path),
                    mime="application/octet-stream",
                    use_container_width=True,
                )
    with restore_col:
        restore_admin_password = st.text_input("Admin password for restore", type="password", key="restore_admin_password")
        restore_file = st.file_uploader("Restore SQLite backup", type=["db"], key="restore_backup_file")
        if restore_file and st.button("RESTORE BACKUP", use_container_width=True):
            if not ADMIN_PASSWORD or not hmac.compare_digest(restore_admin_password, ADMIN_PASSWORD):
                st.error("Administrator authentication failed.")
                st.stop()
            temporary_restore = os.path.join(APP_DIR, ".uploaded-restore.db")
            with open(temporary_restore, "wb") as restore_target:
                restore_target.write(restore_file.getvalue())
            try:
                restored_state = restore_backup(temporary_restore)
                for key, value in restored_state.items():
                    st.session_state[key] = value
                write_excel(st.session_state)
                write_separate_excel_files(st.session_state)
                st.success("Backup restored successfully.")
                st.rerun()
            except ValueError as error:
                st.error(str(error))
            finally:
                if os.path.exists(temporary_restore):
                    os.remove(temporary_restore)
    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as excel_file:
            st.download_button(
                "📊 DOWNLOAD ALL DATA AS ONE EXCEL FILE",
                data=excel_file.read(),
                file_name="ailyn_project_ledger.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    download_materials, download_labor = st.columns(2)
    with download_materials:
        if os.path.exists(MATERIALS_EXCEL_FILE):
            with open(MATERIALS_EXCEL_FILE, "rb") as materials_file:
                st.download_button(
                    "📦 DOWNLOAD MATERIALS EXCEL",
                    data=materials_file.read(),
                    file_name="materials_ledger.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    with download_labor:
        if os.path.exists(LABOR_EXCEL_FILE):
            with open(LABOR_EXCEL_FILE, "rb") as labor_file:
                st.download_button(
                    "👷 DOWNLOAD LABOR EXCEL",
                    data=labor_file.read(),
                    file_name="labor_ledger.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
    if st.button("🗑️ CLEAR RECEIPT HISTORY", use_container_width=True, key="clear_receipt_history"):
        clear_saved_reports()
        st.success("Previous receipt history cleared.")
        st.rerun()
    if st.button("⬅️ BACK TO CONSTRUCTION EXPORT", use_container_width=True):
        set_view("export")
    if st.button("⬅️ BACK TO PAYROLL EXPORT", use_container_width=True):
        set_view("payroll_export")

    for title, report_type in [("🏗️ Construction Receipts", "construction"), ("👷 Payroll Receipts", "payroll")]:
        with st.expander(title, expanded=True):
            saved_reports = list_saved_reports(report_type)
            if not saved_reports:
                st.info(f"No saved {report_type} receipts yet.")
                continue
            for report_path in saved_reports:
                st.markdown(f"- **{report_path.name}**")
                with open(report_path, "r", encoding="utf-8") as handle:
                    report_html = handle.read()
                st.components.v1.html(
                    report_html,
                    height=receipt_preview_height(
                        len(st.session_state.labor_records)
                        if report_type == "payroll"
                        else len(st.session_state.records),
                        row_height=58 if report_type == "payroll" else 35,
                    ),
                    scrolling=True,
                )
                st.download_button(
                    label="📥 DOWNLOAD THIS RECEIPT HTML",
                    data=report_html,
                    file_name=report_path.name,
                    mime="text/html",
                    use_container_width=True,
                    key=f"download_{report_type}_{report_path.name}"
                )
                if st.button("❌ DELETE THIS RECEIPT", key=f"delete_{report_type}_{report_path.name}",
                             use_container_width=True):
                    delete_report_file(report_path)
                    st.success(f"Deleted: {report_path.name}")
                    st.rerun()

elif view == "settings":
    settings = st.session_state.app_settings
    st.markdown("## SETTINGS")
    st.caption("Manage your account, project preferences, notifications, and security.")
    profile_tab, preference_tab, security_tab = st.tabs(["PROFILE", "PREFERENCES", "SECURITY"])
    with profile_tab:
        with st.form("account_profile_form"):
            display_name = st.text_input("Display name", value=settings.get("display_name", ""), placeholder="Your name")
            email = st.text_input("Email address", value=settings.get("email", ""), placeholder="name@example.com")
            if st.form_submit_button("SAVE PROFILE", use_container_width=True):
                if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
                    st.error("Enter a valid email address.")
                else:
                    settings.update({"display_name": display_name.strip(), "email": email.strip()})
                    persist_state()
                    st.success("Profile saved.")
    with preference_tab:
        with st.form("app_preferences_form"):
            dark_mode = st.checkbox("Dark mode", value=bool(st.session_state.dark_mode))
            client_mode = st.checkbox("Client view mode", value=bool(settings.get("client_mode", False)), help="Use this preference when preparing a client-facing project view.")
            email_notifications = st.checkbox("Email notifications", value=bool(settings.get("email_notifications", True)))
            budget_alerts = st.checkbox("Budget alerts", value=bool(settings.get("budget_alerts", True)))
            date_format = st.selectbox("Date format", ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"], index=["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"].index(settings.get("date_format", "%Y-%m-%d")))
            if st.form_submit_button("SAVE PREFERENCES", use_container_width=True):
                st.session_state.dark_mode = dark_mode
                settings.update({"client_mode": client_mode, "email_notifications": email_notifications, "budget_alerts": budget_alerts, "date_format": date_format})
                persist_state()
                st.success("Preferences saved.")
                st.rerun()
    with security_tab:
        st.markdown("### Account access")
        if LOGIN_PASSWORD:
            st.success("Workspace password login is enabled.")
            if st.button("SIGN OUT", use_container_width=True, key="settings_sign_out"):
                st.session_state.authenticated = False
                st.rerun()
        else:
            st.info("Password login is disabled. Set AILYN_LOGIN_PASSWORD to protect this workspace.")
        st.markdown("### Social login readiness")
        google_ready = bool(os.getenv("GOOGLE_CLIENT_ID"))
        facebook_ready = bool(os.getenv("FACEBOOK_APP_ID"))
        st.write(f"Google login: {'Configured' if google_ready else 'Needs OAuth configuration'}")
        st.write(f"Facebook login: {'Configured' if facebook_ready else 'Needs OAuth configuration'}")
        st.caption("Social login requires provider credentials, redirect URLs, and HTTPS. Add those through your deployment secrets; never save client secrets in app data.")
        st.markdown("### Data protection")
        st.write(f"Database backups available: {history_count() > 0}")
        if st.button("CREATE BACKUP", use_container_width=True, key="settings_backup"):
            backup_path = create_backup()
            with open(backup_path, "rb") as backup_file:
                st.download_button("DOWNLOAD BACKUP", backup_file.read(), file_name=os.path.basename(backup_path), mime="application/octet-stream", use_container_width=True, key="settings_download_backup")

elif view == "photo_scanner":
    st.markdown("## PHOTO STUDIO")
    st.caption("Capture a receipt or project update. The app will read and organize the photo for you.")
    studio_capture, studio_status = st.columns([1.25, 0.75])
    with studio_capture:
        st.markdown("### Camera")
        st.markdown("Use your device camera to create a clear project record.")
        if st.button("📷 OPEN CAMERA", use_container_width=True, key="open_photo_studio"):
            st.session_state.scanner_open = True
            st.rerun()
        if st.session_state.scanner_open:
            photo_camera_dialog()
    with studio_status:
        st.markdown("### Studio status")
        st.metric("Saved photos", len(st.session_state.get("scanner_photos", [])))
        if st.session_state.get("scanned_photo_bytes"):
            st.success("Photo ready for review")
            st.caption("Review the scan, save the file, or retake the photo from the camera window.")
        else:
            st.info("No photo captured yet")
        if st.button("OPEN PHOTO GALLERY", use_container_width=True, key="studio_gallery"):
            set_view("project_tools")
    if st.button("BACK TO DASHBOARD", key="studio_back"):
        set_view("home")

elif view == "project_tools":
    st.subheader("PROJECT TOOLS")
    st.caption("Organize project evidence, find records quickly, and prepare focused reports.")
    tools_gallery, tools_search, tools_cleanup = st.tabs(["PHOTO GALLERY", "SEARCH & REPORTS", "DATA CLEANUP"])

    with tools_gallery:
        uploaded_photos = st.file_uploader("Add project photos", type=["jpg", "jpeg", "png", "webp"], accept_multiple_files=True, key="project_photo_uploads")
        upload_tag = st.selectbox("Photo category", ["General", "Before", "After", "Framing", "Electrical", "Plumbing", "Painting", "Inspection"], key="project_photo_tag")
        if uploaded_photos and st.button("SAVE ALL PHOTOS", use_container_width=True, key="save_project_photos"):
            existing_hashes = {photo.get("hash") for photo in st.session_state.scanner_photos}
            saved_count = 0
            for uploaded in uploaded_photos:
                photo_bytes, photo_mime = normalize_photo_bytes(uploaded.getvalue(), uploaded.type or "image/jpeg")
                photo_hash = hashlib.sha256(photo_bytes).hexdigest()
                if photo_hash in existing_hashes:
                    continue
                photo_id = str(uuid.uuid4())
                relative_path = save_scanner_photo(photo_bytes, photo_mime, photo_id)
                st.session_state.scanner_photos.append({"id": photo_id, "hash": photo_hash, "file": relative_path, "tag": upload_tag, "saved_at": manila_now().isoformat()})
                existing_hashes.add(photo_hash)
                saved_count += 1
            persist_state()
            st.success(f"Saved {saved_count} new photo(s).")
        photos = st.session_state.get("scanner_photos", [])
        if not photos:
            st.info("No saved project photos yet.")
        else:
            selected_tag = st.selectbox("Filter photos", ["All"] + sorted({photo.get("tag", "General") for photo in photos}), key="gallery_filter")
            visible_photos = [photo for photo in photos if selected_tag == "All" or photo.get("tag", "General") == selected_tag]
            gallery_columns = st.columns(4)
            for index, photo in enumerate(visible_photos):
                photo_path = os.path.abspath(os.path.join(APP_DIR, photo.get("file", "")))
                if not os.path.isfile(photo_path):
                    continue
                with gallery_columns[index % 4]:
                    with open(photo_path, "rb") as image_file:
                        image_bytes = image_file.read()
                    st.image(image_bytes, use_container_width=True)
                    st.caption(f"{photo.get('tag', 'General')} | {photo.get('saved_at', '')[:10]}")
                    st.download_button("DOWNLOAD", image_bytes, file_name=os.path.basename(photo_path), key=f"download_photo_{photo['id']}", use_container_width=True)
            if len(visible_photos) >= 2:
                st.markdown("#### BEFORE / AFTER COMPARISON")
                photo_options = {f"{photo.get('tag', 'General')} | {photo.get('saved_at', '')[:10]} | {photo['id'][:8]}": photo for photo in visible_photos}
                compare_left, compare_right = st.columns(2)
                with compare_left:
                    left_label = st.selectbox("Before photo", list(photo_options), key="compare_left")
                with compare_right:
                    right_label = st.selectbox("After photo", list(photo_options), index=min(1, len(photo_options) - 1), key="compare_right")
                comparison_columns = st.columns(2)
                for column, label in zip(comparison_columns, (left_label, right_label)):
                    comparison_path = os.path.join(APP_DIR, photo_options[label].get("file", ""))
                    if os.path.isfile(comparison_path):
                        with column:
                            st.image(comparison_path, use_container_width=True)

    with tools_search:
        query = st.text_input("Search materials, labor, payroll, and tasks", key="global_search").strip().lower()
        all_records = searchable_records(st.session_state)
        all_records += [{"category": "Task", **task} for task in st.session_state.get("planner_tasks", [])]
        results = [record for record in all_records if not query or query in " ".join(str(value) for value in record.values()).lower()]
        st.metric("Matching records", len(results))
        if results:
            st.dataframe(results, use_container_width=True, hide_index=True)
        else:
            st.info("No matching records.")
        st.markdown("#### CUSTOM REPORT")
        report_type = st.selectbox("Report data", ["All records", "Materials only", "Expenses only"], key="custom_report_type")
        report_records = [record for record in all_records if report_type == "All records" or record.get("category") == report_type.removesuffix(" only")]
        report_records = [{
            "type": "material" if record.get("category") == "Material" else "expense",
            "date": record.get("date", record.get("month", "")),
            "qty": record.get("qty", 1),
            "name": record.get("name", record.get("item", record.get("description", ""))),
            "price": record.get("price", record.get("amount", record.get("net", 0))),
            "delivery": record.get("delivery", 0),
            "amount": record.get("amount", record.get("price", record.get("net", 0))),
        } for record in report_records if record.get("category") != "Task"]
        report_html = build_html_report(report_records, st.session_state.budget, custom_title="CUSTOM PROJECT REPORT")
        st.download_button("DOWNLOAD CUSTOM REPORT", report_html, file_name="custom_project_report.html", mime="text/html", use_container_width=True)

    with tools_cleanup:
        duplicate_groups = find_duplicate_records(searchable_records(st.session_state))
        st.metric("Duplicate groups", len(duplicate_groups))
        for group in duplicate_groups:
            st.warning("Duplicate: " + " | ".join(str(item.get("name", item.get("item", item.get("description", "record")))) for item in group))
        if duplicate_groups and st.button("REMOVE DUPLICATE RECORDS", use_container_width=True, key="remove_duplicates"):
            seen = set()
            for key in ("records", "labor_records", "payroll_expenses"):
                kept = []
                for record in st.session_state.get(key, []):
                    signature = (record.get("type", key), str(record.get("name", record.get("item", record.get("description", "")))).strip().lower(), round(float(record.get("amount", record.get("price", record.get("net", 0))) or 0), 2), record.get("date", record.get("month", "")))
                    if signature not in seen:
                        seen.add(signature)
                        kept.append(record)
                st.session_state[key] = kept
            persist_state()
            st.success("Duplicate records removed; the first copy was kept.")
            st.rerun()
        st.markdown("#### DISPLAY")
        dark_mode = st.toggle("Dark mode", value=st.session_state.dark_mode, key="dark_mode_toggle")
        if dark_mode != st.session_state.dark_mode:
            st.session_state.dark_mode = dark_mode
            persist_state()
            st.rerun()

elif view == "update":
    st.markdown("## Upgrade Center")
    st.caption("Administrator-only signed release installation. The current app is backed up first.")
    admin_password = st.text_input("Administrator password", type="password", key="admin_update_password")
    uploaded_upgrade = st.file_uploader("Choose signed Python upgrade", type=["py"], key="upgrade_file")
    signature = st.text_input("Release SHA-256 HMAC signature", key="upgrade_signature")
    confirm_upgrade = st.checkbox("I have reviewed this signed release and want to install it.")
    if st.button("INSTALL SIGNED UPGRADE", use_container_width=True,
                 disabled=not (admin_password and uploaded_upgrade and signature and confirm_upgrade)):
        try:
            if not ADMIN_PASSWORD or not hmac.compare_digest(admin_password, ADMIN_PASSWORD):
                raise ValueError("Administrator authentication failed.")
            backup_path = install_update(uploaded_upgrade, signature)
            st.success(f"Upgrade installed. Backup created at {os.path.basename(backup_path)}.")
            st.warning("Restart the Streamlit app to load the new version.")
        except ValueError as error:
            st.error(str(error))
    st.divider()
    st.subheader("Backups")
    backup_dir = os.path.join(APP_DIR, "backups")
    backup_names = sorted(os.listdir(backup_dir), reverse=True) if os.path.isdir(backup_dir) else []
    if backup_names:
        st.dataframe([{"Backup": name} for name in backup_names[:10]], use_container_width=True, hide_index=True)
    else:
        st.caption("No upgrades have been installed yet.")

else:
    st.info("Welcome to Ailyn Project Management System. Use the command sidebar to navigate.")
